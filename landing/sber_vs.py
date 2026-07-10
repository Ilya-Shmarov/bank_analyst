# -*- coding: utf-8 -*-
"""Static bank-vs-bank comparison landing generated from the Excel summary.

Интеракция: пользователь выбирает Банк 1 → уровень пакета этого банка,
затем Банк 2 → уровень пакета; сравнение появляется на том же экране без
прокрутки страницы. Внутри сравнения — своя скроллируемая область с
зафиксированной шапкой выбранных уровней.

Терминология UI — «уровень пакета». Внутренний ключ данных `tier`
(history.json, changelog, колонка «Тир» в Excel) не переименовывается:
его читают Excel-отчёт и этот модуль как контракт данных.
"""

import html
import json
import re
from pathlib import Path

from openpyxl import load_workbook

from scanner.scoring import SCORERS, WEIGHTS
from scanner.sources import NOT_FOUND

SUMMARY_SHEET = "Сводная"
INTL_SEGMENT = "digital-first (межд.)"

# Заголовки колонок Excel-листа «Сводная» — контракт с report/excel_writer.py.
# «Тир» — имя колонки данных; в UI лендинга показывается «уровень пакета».
BASE_COLUMNS = {
    "segment": "Сегмент капитала",
    "bank": "Банк",
    "tier": "Тир",
    "scan_date": "Дата скана",
    "sources_ok": "Источников OK",
    "score": "Итоговый балл (0–5)",
    "divergent": "Расхождение источников",
}

FIELD_COLUMNS = {
    "entry_conditions": "Условия входа / поддержания уровня",
    "service_cost": "Стоимость обслуживания",
    "lounge_access": "Бизнес-залы (визиты, спутники)",
    "concierge": "Консьерж-сервис",
    "cashback": "Кэшбэк (ставка, категории, механика)",
    "deposits": "Спецусловия по вкладам / накопительным счетам",
    "insurance": "Страхование (мед., ВЗР)",
    "auto": "Автоуслуги",
    "taxi_restaurants": "Такси и рестораны (компенсации)",
    "ecosystem": "Экосистемные привилегии (доставка, подписки)",
}

FIELD_LABELS = {
    "entry_conditions": "Условия входа",
    "service_cost": "Стоимость обслуживания",
    "lounge_access": "Бизнес-залы",
    "concierge": "Консьерж",
    "cashback": "Кэшбэк",
    "deposits": "Вклады",
    "insurance": "Страхование",
    "auto": "Авто",
    "taxi_restaurants": "Такси и рестораны",
    "ecosystem": "Экосистема",
}

# Порядок атрибутов в таблице сравнения
COMPARE_FIELDS = (
    "entry_conditions",
    "service_cost",
    "lounge_access",
    "cashback",
    "deposits",
    "concierge",
    "insurance",
    "taxi_restaurants",
    "ecosystem",
    "auto",
)

# Текст пояснения итогового балла — перенос логики листа «Методика оценки»
# Excel-отчёта (scanner/scoring.py: METHODOLOGY_TEXT + WEIGHTS), не пересказ.
def _methodology_text() -> str:
    w = WEIGHTS
    return (
        "Итоговый балл пакета — сумма баллов категорий (0–5), умноженных "
        "на вес категории; шкала 0–5. Балл категории получается из метрики, "
        "извлечённой из собранных данных, по пороговым таблицам листа "
        "«Методика оценки» Excel-отчёта — всё воспроизводимо вручную. "
        f"Веса различаются: бизнес-залы и кэшбэк — по {w['lounge_access']:.2f}, "
        f"вклады — {w['deposits']:.2f}, такси и рестораны, страхование, "
        f"консьерж и экосистема — по {w['concierge']:.2f}, "
        f"авто — {w['auto']:.2f}. Если по категории данных нет, она получает "
        "0 — отсутствие данных снижает балл, а не трактуется в пользу банка."
    )


def build_sber_vs_landing(workbook_path: Path, output_path: Path) -> dict:
    """Build the static bank comparison landing page."""
    rows = load_summary_rows(workbook_path)
    banks = build_payload(rows)
    html_text = render_html(banks, rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(html_text, encoding="utf-8")
    return {
        "output": str(output_path),
        "banks": len(banks),
        "levels": sum(len(b["levels"]) for b in banks),
    }


def load_summary_rows(workbook_path: Path) -> list[dict]:
    """Read normalized banking rows from the workbook summary sheet."""
    if not workbook_path.exists():
        raise FileNotFoundError(
            f"Excel report not found: {workbook_path}. Run --scan-all first."
        )

    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    if SUMMARY_SHEET not in wb.sheetnames:
        raise ValueError(f'Sheet "{SUMMARY_SHEET}" not found in {workbook_path}.')

    ws = wb[SUMMARY_SHEET]
    headers = {
        _clean(cell.value): idx
        for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), start=1)
    }
    _require_headers(headers, [*BASE_COLUMNS.values(), *FIELD_COLUMNS.values()])

    rows = []
    current_segment = ""
    for cells in ws.iter_rows(min_row=2, values_only=True):
        raw_segment = _clean(cells[headers[BASE_COLUMNS["segment"]] - 1])
        bank = _clean(cells[headers[BASE_COLUMNS["bank"]] - 1])
        tier = _clean(cells[headers[BASE_COLUMNS["tier"]] - 1])
        if raw_segment and not bank:
            current_segment = raw_segment
            continue
        if not bank or not tier:
            continue

        row = {
            key: _clean(cells[headers[column] - 1])
            for key, column in BASE_COLUMNS.items()
        }
        row["segment"] = row["segment"] or current_segment
        row["score"] = _parse_float(row["score"])
        row["fields"] = {
            key: _clean(cells[headers[column] - 1])
            for key, column in FIELD_COLUMNS.items()
        }
        if row["segment"] and row["segment"] != INTL_SEGMENT:
            rows.append(row)
    return rows


def build_payload(rows: list[dict]) -> list[dict]:
    """Group rows into banks → levels with pre-computed compare attributes."""
    by_bank = {}
    for row in rows:
        by_bank.setdefault(row["bank"], []).append(row)

    banks = []
    # Сбер первым (база сравнения), остальные по алфавиту
    order = sorted(by_bank, key=lambda name: (name != "Сбер", name.lower()))
    for name in order:
        levels = []
        for row in by_bank[name]:
            attrs = []
            for field in COMPARE_FIELDS:
                metric = _attr_metric(field, row)
                attrs.append({
                    "id": field,
                    "label": FIELD_LABELS[field],
                    "value": metric["value"],
                    # score: число для подсветки сильной стороны, null — не сравниваем
                    "score": metric["score"],
                    "note": metric["note"],
                })
            levels.append({
                "tier": row["tier"],  # значение поля данных; в UI — уровень пакета
                "segment": row["segment"],
                "score": row["score"],
                "score_str": _format_score(row["score"]),
                "scan_date": (row.get("scan_date") or "")[:10],
                "attrs": attrs,
            })
        banks.append({"bank": name, "levels": levels})
    return banks


def _attr_metric(field: str, row: dict) -> dict:
    """Attribute display value + comparable score for one level."""
    raw = row["fields"].get(field, "")
    if field == "entry_conditions":
        summary = _condition_summary(raw)
        return {"value": summary or "нет данных", "score": None,
                "note": _shorten(raw, 260)}
    if field == "service_cost":
        cost_info = _service_cost_summary(row)
        cost = _monthly_rub_cost(raw)
        return {"value": cost_info["display"],
                "score": -cost if cost is not None else None,
                "note": _shorten(raw, 260)}

    if _is_missing(raw):
        return {"value": "нет данных", "score": 0, "note": ""}
    if raw.strip().startswith(("—", "-")):
        return {"value": "не предусмотрено", "score": 0, "note": ""}
    try:
        metric, score = SCORERS[field](raw)
    except Exception:  # noqa: BLE001 — шумный текст Excel не роняет лендинг
        metric, score = ("есть, детали не выделены", 1) if _has_benefit(raw) else ("нет", 0)
    value = _display_text(metric)
    if isinstance(score, (int, float)):
        value = f"{value} ({_format_metric_score(score)})"
    return {"value": value, "score": score, "note": _shorten(raw, 260)}


# ---------- рендер ----------

def render_html(banks: list[dict], rows: list[dict]) -> str:
    scan_dates = sorted({r["scan_date"][:10] for r in rows if r.get("scan_date")})
    latest_scan = scan_dates[-1] if scan_dates else "нет данных"
    total_levels = sum(len(b["levels"]) for b in banks)
    payload = json.dumps(banks, ensure_ascii=False).replace("</", "<\\/")
    bank_chips = _render_bank_chips(banks)

    return f"""<!DOCTYPE html>
<html lang="ru">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Сравнение премиальных пакетов банков</title>
  <style>{_CSS}</style>
</head>
<body>
  <main class="page">
    <section class="hero">
      <p class="eyebrow">Премиальный банкинг РФ</p>
      <h1>Сравнение уровней пакетов</h1>
      <p class="lead">Выберите два банка и уровень пакета у каждого —
      сравнение по итоговому баллу и ключевым привилегиям появится сразу,
      без прокрутки страницы.</p>
      <div class="stats">
        <div><b>{len(banks)}</b><span>банков</span></div>
        <div><b>{total_levels}</b><span>уровней пакетов</span></div>
        <div><b>{_esc(latest_scan)}</b><span>дата данных</span></div>
      </div>
    </section>

    <section class="pickers">
      <div class="picker" data-side="a">
        <h2>Банк 1</h2>
        <div class="chip-row banks">{bank_chips}</div>
        <h3 class="lvl-title" hidden>Уровень пакета</h3>
        <div class="chip-row levels"></div>
      </div>
      <div class="picker" data-side="b">
        <h2>Банк 2</h2>
        <div class="chip-row banks">{bank_chips}</div>
        <h3 class="lvl-title" hidden>Уровень пакета</h3>
        <div class="chip-row levels"></div>
      </div>
    </section>
    <p id="js-warning" class="js-warning">Если банки не выбираются, файл открыт
    во встроенном просмотрщике без JavaScript. Нажмите «Поделиться» → «Открыть
    в Safari» или откройте этот HTML в Chrome.</p>

    <section id="compare" hidden>
      <div class="cmp-head">
        <div class="cmp-col" data-head="a"></div>
        <div class="cmp-col" data-head="b"></div>
      </div>
      <details class="method">
        <summary>Как считается итоговый балл 0–5</summary>
        <p>{html.escape(_methodology_text())}</p>
      </details>
      <div class="cmp-scroll">
        <table class="cmp-table">
          <colgroup>
            <col class="attr"><col class="side"><col class="side">
          </colgroup>
          <thead>
            <tr><th>Атрибут</th><th data-th="a"></th><th data-th="b"></th></tr>
          </thead>
          <tbody></tbody>
        </table>
      </div>
    </section>
    <p id="hint" class="hint">Выберите банк и уровень пакета слева и справа —
    сравнение появится здесь.</p>

    <footer class="footer">
      <p>Данные — из Excel-отчёта сканера (лист «Сводная»), у каждого значения
      в отчёте зафиксирован источник и дата проверки. Дата данных:
      {_esc(latest_scan)}.</p>
    </footer>
  </main>
  <script id="data" type="application/json">{payload}</script>
  <script>{_JS}</script>
</body>
</html>"""


# ---------- значения и форматирование ----------

def _render_bank_chips(banks: list[dict]) -> str:
    return "".join(
        f'<button type="button" class="chip" data-bank-index="{idx}">'
        f'{_esc(bank["bank"])}</button>'
        for idx, bank in enumerate(banks)
    )

def _require_headers(headers: dict, required: list[str]):
    missing = [header for header in required if header not in headers]
    if missing:
        raise ValueError(
            f"Missing required columns in sheet \"{SUMMARY_SHEET}\": "
            + ", ".join(missing)
        )


def _service_cost_summary(row) -> dict:
    raw = row["fields"].get("service_cost", "")
    entry = row["fields"].get("entry_conditions", "")
    combined_low = f"{raw} {entry}".lower()
    parts = []
    if "бесплат" in combined_low or re.search(r"\b0\s*(?:₽|руб)", combined_low,
                                              flags=re.IGNORECASE):
        parts.append("бесплатно при выполнении условий")
    cost = _monthly_rub_cost(raw)
    if cost is not None and cost > 0 and not parts and entry and not _is_missing(entry):
        parts.append("бесплатно при выполнении условий")
    if cost is not None and cost > 0:
        parts.append(f"{_format_rub(cost)} ₽ в месяц")
    if not parts and _is_missing(raw):
        parts.append("стоимость не указана")
    if not parts and raw:
        parts.append(_shorten(raw, 140))
    if not parts:
        parts.append("нет данных")
    parts[0] = parts[0][0].upper() + parts[0][1:]
    return {"display": " или ".join(parts)}


def _condition_summary(value: str, limit: int = 5) -> str:
    text = _public_text(value)
    if not text or _is_missing(text):
        return ""
    text = re.sub(r"\s*;\s*", " | ", text)
    text = re.sub(r"\s+\|\s+или\s+\|", " | ", text)
    text = re.sub(r"\|\s*или\s*\|", "|", text)
    text = re.sub(r"\s+", " ", text).strip(" |;")
    parts = []
    for part in re.split(r"\s*\|\s*", text):
        cleaned = part.strip(" ;")
        if not cleaned:
            continue
        if cleaned.lower().startswith("или "):
            cleaned = cleaned[4:].strip()
        low = cleaned.lower()
        if re.search(r"\d[\d\s]*\s*₽\s*в\s*мес", low):
            continue
        if "последний календарный день" in low:
            continue
        if "среднемесячный остаток" in low and parts:
            continue
        if cleaned and cleaned not in parts:
            parts.append(cleaned)
    if not parts:
        return text
    shown = parts[:limit]
    summary = "; ".join(shown)
    if len(parts) > limit:
        summary += f"; ещё {len(parts) - limit}"
    return summary


def _format_metric_score(score) -> str:
    if score is None:
        return "нет данных"
    if isinstance(score, float) and not score.is_integer():
        return f"{score:.2f}/5"
    return f"{int(score)}/5"


def _clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_float(value):
    if value in ("", None):
        return None
    try:
        return float(str(value).replace(",", "."))
    except ValueError:
        return None


def _monthly_rub_cost(value: str):
    if not value:
        return None
    normalized = value.replace("\xa0", " ")
    match = re.search(
        r"(\d[\d\s.,]*)\s*(?:₽|руб)[^\n;]{0,20}(?:мес|месяц)",
        normalized,
        flags=re.IGNORECASE,
    )
    if not match:
        return None
    digits = re.sub(r"[^\d.,]", "", match.group(1)).replace(",", ".")
    try:
        return float(digits)
    except ValueError:
        return None


def _is_missing(value: str) -> bool:
    text = (value or "").strip()
    return not text or text == NOT_FOUND or "не найдено" in text.lower()


def _has_benefit(value: str) -> bool:
    text = (value or "").strip()
    if _is_missing(text):
        return False
    low = text.lower()
    if low.startswith(("—", "-")):
        return False
    if low.startswith("нет —") or low.startswith("нет,"):
        return False
    return True


def _format_score(value) -> str:
    return "нет данных" if value is None else f"{value:.2f}"


def _format_rub(value) -> str:
    try:
        amount = int(float(value))
    except (TypeError, ValueError):
        return str(value)
    return f"{amount:,}".replace(",", " ")


def _shorten(value: str, limit: int = 120) -> str:
    text = " ".join(_public_text(value or NOT_FOUND).split())
    return text if len(text) <= limit else text[:limit - 1].rstrip() + "…"


def _esc(value) -> str:
    return html.escape(_display_text(value))


def _public_text(value) -> str:
    text = str(value or "")
    text = re.sub(r"\s*\[(?:источник|проверено|прим\.)[^\]]*\]", "", text,
                  flags=re.IGNORECASE)
    text = re.sub(r"\s*\[[^\]]*(?:источник|проверено|первоисточник)[^\]]*\]",
                  "", text, flags=re.IGNORECASE)
    text = text.replace(" ;", ";").replace("| |", "|")
    return " ".join(text.split())


def _display_text(value) -> str:
    text = _public_text(value).replace("₽/мес", "₽ в мес")
    text = re.sub(r"(\d+)\s+визит\(ов\)/мес", _visit_text, text)
    text = re.sub(r"(\d+)\s+компенсаций/мес суммарно", _compensation_text, text)
    text = re.sub(r"(\d+)\s+опций/подписок упомянуто", _option_text, text)
    replacements = {
        "Excel": "таблица",
        "excel": "таблица",
        "автоматически": "",
        "распознанная метрика": "метрика",
        "распознанный балл": "балл",
        "распознанного": "",
        "распознанной": "",
        "распознано": "выделено",
        "распознан": "выделен",
        "исходные фрагменты": "детали",
        "исходный фрагмент": "детали",
        "безусловное": "",
        "проседает": "ниже",
        "выигрывает": "сильнее",
        "дешевле": "ниже по цене",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return " ".join(text.split())


def _visit_text(match) -> str:
    count = int(match.group(1))
    noun = _ru_plural(count, "визит", "визита", "визитов")
    return f"{count} {noun} в месяц"


def _compensation_text(match) -> str:
    count = int(match.group(1))
    noun = _ru_plural(count, "компенсация", "компенсации", "компенсаций")
    return f"{count} {noun} в месяц"


def _option_text(match) -> str:
    count = int(match.group(1))
    option = _ru_plural(count, "опция", "опции", "опций")
    subscription = _ru_plural(count, "подписка", "подписки", "подписок")
    return f"{count} {option} или {subscription}"


def _ru_plural(count: int, one: str, few: str, many: str) -> str:
    if count % 10 == 1 and count % 100 != 11:
        return one
    elif count % 10 in (2, 3, 4) and count % 100 not in (12, 13, 14):
        return few
    return many


# Токен-сет единый с лендингом отзывов (landing/premium_reviews.py):
# белый фон, зелёный акцент, терракот — только для негативного сигнала
# (здесь — меньший итоговый балл в шапке сравнения).
_CSS = """
:root {
  --bg: #FFFFFF;
  --card: #FAFAF8;
  --ink: #1A1D1B;
  --muted: #64746b;
  --line: #E7E5DE;
  --green: #188f4f;
  --green-soft: #188f4f1a;
  --neg: #B3492F;
}
* { box-sizing: border-box; }
html { -webkit-text-size-adjust: 100%; }
body {
  margin: 0; background: var(--bg); color: var(--ink);
  font: 15px/1.55 -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
  -webkit-tap-highlight-color: rgba(24, 143, 79, 0.18);
}
button, summary { touch-action: manipulation; }
button { -webkit-appearance: none; appearance: none; }
.page { max-width: 1160px; margin: 0 auto; padding: 36px 18px 56px; }
.hero { padding: 18px 0 24px; border-bottom: 1px solid var(--line); }
.eyebrow { margin: 0 0 8px; color: var(--green); font-size: 13px;
  font-weight: 700; text-transform: uppercase; }
h1 { margin: 0; font-size: 42px; line-height: 1.08; }
.lead { max-width: 780px; margin: 14px 0 0; color: var(--muted); font-size: 17px; }
.stats { display: flex; flex-wrap: wrap; gap: 10px; margin-top: 22px; }
.stats div { min-width: 150px; background: var(--card);
  border: 1px solid var(--line); border-radius: 8px; padding: 12px 14px; }
.stats b { display: block; font-size: 24px; color: var(--green);
  font-family: ui-monospace, "SF Mono", Menlo, monospace; }
.stats span { color: var(--muted); font-size: 13px; }
.pickers { display: grid; grid-template-columns: 1fr 1fr; gap: 16px;
  margin-top: 22px; }
.picker { background: var(--card); border: 1px solid var(--line);
  border-radius: 8px; min-width: 0; padding: 14px 16px; }
.picker h2 { margin: 0 0 10px; font-size: 16px; }
.picker h3 { margin: 12px 0 8px; font-size: 12px; color: var(--muted);
  text-transform: uppercase; }
.chip-row { display: flex; flex-wrap: wrap; gap: 6px; }
.chip { border: 1px solid var(--line); background: var(--bg); color: var(--ink);
  border-radius: 999px; min-height: 44px; padding: 10px 14px; font-size: 14px;
  line-height: 1.2; cursor: pointer; font-family: inherit; }
.chip:hover { border-color: var(--green); color: var(--green); }
.chip.active { background: var(--green); border-color: var(--green); color: #fff; }
.hint { margin: 18px 0 0; color: var(--muted); }
.js-warning { margin: 14px 0 0; padding: 12px 14px; border: 1px solid var(--line);
  border-radius: 8px; background: #fff8e8; color: #6f5a25; font-size: 14px; }
.js-ready .js-warning { display: none; }
#compare { margin-top: 22px; }
.cmp-head { display: grid; grid-template-columns: 1fr 1fr; gap: 16px; }
.cmp-col { background: var(--card); border: 1px solid var(--line);
  border-top: 3px solid var(--green); border-radius: 8px; min-width: 0;
  padding: 14px 16px; }
.cmp-col .seg { margin: 0; color: var(--muted); font-size: 12px; }
.cmp-col h2 { margin: 2px 0 6px; font-size: 20px; }
.cmp-col .sc { font-family: ui-monospace, "SF Mono", Menlo, monospace;
  font-size: 26px; color: var(--green); font-weight: 700; }
.cmp-col .sc.lower { color: var(--neg); }
.cmp-col .sc small { font-size: 12px; color: var(--muted); font-weight: 400; }
.method { margin: 12px 0 0; color: var(--muted); font-size: 13px; }
.method summary { cursor: pointer; font-weight: 700; color: var(--ink); }
.method p { margin: 6px 0 0; max-width: 900px; }
.cmp-scroll { margin-top: 12px; max-height: 62vh; overflow-y: auto;
  border: 1px solid var(--line); border-radius: 8px; }
.cmp-table { width: 100%; border-collapse: collapse; table-layout: fixed; }
.cmp-table col.attr { width: 18%; }
.cmp-table col.side { width: 41%; }
.cmp-table th, .cmp-table td { text-align: left; padding: 10px 12px;
  border-bottom: 1px solid var(--line); vertical-align: top; font-size: 14px;
  overflow-wrap: anywhere; }
.cmp-table thead th { position: sticky; top: 0; background: var(--card);
  color: var(--muted); font-size: 12px; text-transform: uppercase; z-index: 1;
  box-shadow: 0 1px 0 var(--line); }
.cmp-table td:first-child { color: var(--muted); font-size: 13px;
  white-space: nowrap; }
.cmp-table td.win { box-shadow: inset 3px 0 0 var(--green);
  background: var(--green-soft); }
.cmp-table td .tag { display: inline-block; margin-left: 6px;
  background: var(--green-soft); color: var(--green); border-radius: 999px;
  padding: 1px 8px; font-size: 11px; font-weight: 700; }
.footer { margin-top: 40px; padding-top: 18px; border-top: 1px solid var(--line);
  color: var(--muted); font-size: 13px; }
@media (max-width: 820px) {
  .page { padding: 24px 14px 42px; }
  h1 { font-size: 32px; }
  .stats div { flex: 1 1 138px; min-width: 0; }
  .pickers, .cmp-head { grid-template-columns: 1fr; }
  .picker { padding: 14px; }
  .chip-row { gap: 8px; }
  .chip { flex: 1 1 auto; justify-content: center; min-width: min(46%, 220px); }
  .cmp-scroll { max-height: none; overflow: visible; border: 0; border-radius: 0; }
  .cmp-table, .cmp-table colgroup, .cmp-table tbody, .cmp-table tr,
  .cmp-table td { display: block; width: 100%; }
  .cmp-table thead { display: none; }
  .cmp-table tr { margin-bottom: 12px; border: 1px solid var(--line);
    border-radius: 8px; background: var(--card); overflow: hidden; }
  .cmp-table td { border-bottom: 1px solid var(--line); padding: 11px 12px;
    white-space: normal; }
  .cmp-table td:first-child { background: var(--bg); font-weight: 700;
    color: var(--ink); font-size: 14px; }
  .cmp-table td:last-child { border-bottom: 0; }
  .cmp-table td[data-label]::before { content: attr(data-label); display: block;
    margin-bottom: 4px; color: var(--muted); font-size: 12px; font-weight: 700; }
  .cmp-table td.win { box-shadow: inset 3px 0 0 var(--green); }
}
"""

_JS = """
const DATA = JSON.parse(document.getElementById('data').textContent);
const state = { a: { bank: null, level: null }, b: { bank: null, level: null } };
document.documentElement.classList.add('js-ready');

function el(tag, cls, text) {
  const node = document.createElement(tag);
  if (cls) node.className = cls;
  if (text !== undefined) node.textContent = text;
  return node;
}

function renderBanks(side) {
  const picker = document.querySelector(`.picker[data-side="${side}"]`);
  const row = picker.querySelector('.banks');
  if (!row.children.length) {
    DATA.forEach((bank, i) => {
      const chip = el('button', 'chip', bank.bank);
      chip.type = 'button';
      chip.dataset.bankIndex = i;
      row.appendChild(chip);
    });
  }
  row.querySelectorAll('.chip').forEach((chip) => {
    const i = Number(chip.dataset.bankIndex);
    if (state[side].bank === i) chip.classList.add('active');
    else chip.classList.remove('active');
    chip.onclick = () => {
      state[side].bank = i;
      state[side].level = null;
      renderBanks(side);
      renderLevels(side);
      renderCompare();
    };
  });
}

function renderLevels(side) {
  const picker = document.querySelector(`.picker[data-side="${side}"]`);
  const title = picker.querySelector('.lvl-title');
  const row = picker.querySelector('.levels');
  row.innerHTML = '';
  const bankIdx = state[side].bank;
  title.hidden = bankIdx === null;
  if (bankIdx === null) return;
  DATA[bankIdx].levels.forEach((lvl, i) => {
    const chip = el('button', 'chip', lvl.tier);
    chip.type = 'button';
    if (state[side].level === i) chip.classList.add('active');
    chip.onclick = () => {
      state[side].level = i;
      renderLevels(side);
      renderCompare();
    };
    row.appendChild(chip);
  });
}

function selected(side) {
  const s = state[side];
  if (s.bank === null || s.level === null) return null;
  return { bank: DATA[s.bank].bank, ...DATA[s.bank].levels[s.level] };
}

function renderHead(node, item, other) {
  node.innerHTML = '';
  node.appendChild(el('p', 'seg', item.segment));
  node.appendChild(el('h2', '', item.bank + ' — ' + item.tier));
  const sc = el('div', 'sc', item.score_str);
  if (item.score !== null && other && other.score !== null
      && item.score < other.score) {
    sc.classList.add('lower');
  }
  const label = el('small', '', ' итоговый балл 0–5');
  sc.appendChild(label);
  node.appendChild(sc);
}

function renderCompare() {
  const a = selected('a'), b = selected('b');
  const cmp = document.getElementById('compare');
  const hint = document.getElementById('hint');
  if (!a || !b) { cmp.hidden = true; hint.hidden = false; return; }
  cmp.hidden = false; hint.hidden = true;

  renderHead(cmp.querySelector('[data-head="a"]'), a, b);
  renderHead(cmp.querySelector('[data-head="b"]'), b, a);
  cmp.querySelector('[data-th="a"]').textContent = a.bank + ' — ' + a.tier;
  cmp.querySelector('[data-th="b"]').textContent = b.bank + ' — ' + b.tier;

  const tbody = cmp.querySelector('tbody');
  tbody.innerHTML = '';
  a.attrs.forEach((attrA, i) => {
    const attrB = b.attrs[i];
    const tr = el('tr');
    tr.appendChild(el('td', '', attrA.label));
    const tdA = el('td', '', attrA.value);
    const tdB = el('td', '', attrB.value);
    tdA.dataset.label = a.bank + ' — ' + a.tier;
    tdB.dataset.label = b.bank + ' — ' + b.tier;
    if (attrA.note) tdA.title = attrA.note;
    if (attrB.note) tdB.title = attrB.note;
    if (attrA.score !== null && attrB.score !== null
        && attrA.score !== attrB.score) {
      const winner = attrA.score > attrB.score ? tdA : tdB;
      winner.classList.add('win');
      winner.appendChild(el('span', 'tag', 'сильнее'));
    }
    tr.appendChild(tdA);
    tr.appendChild(tdB);
    tbody.appendChild(tr);
  });
  cmp.scrollIntoView({ behavior: 'smooth', block: 'start' });
}

renderBanks('a');
renderBanks('b');
"""
