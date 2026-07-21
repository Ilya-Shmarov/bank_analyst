import json
import re
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from landing.sber_vs import (
    build_sber_vs_landing,
    _compensation_evaluation,
    _condition_summary,
    _service_evaluation,
    _service_cost_summary,
)
from report.json_writer import build_comparison_json
from scanner.benefits import build_other_benefits
from scanner.curated import curated_for
from scanner.formatting import (
    format_natural_list,
    format_list,
    make_complete_summary,
    normalize_user_text,
    normalize_source_text,
    split_summary_and_details,
    validate_user_visible_text,
)
from scanner.diff import load_history
from scanner.parse import parse_source
from scanner.scoring import score_tier
from scanner.sources import BANKS, NOT_FOUND, PRIORITY_SOURCE_URLS, tier_sources
from report.excel_writer import write_report
from report.json_writer import write_comparison_json


def merged(value):
    return {"value": value}


class PremiumStructuredTests(unittest.TestCase):
    USER_VISIBLE_SHEETS = [
        "Сводная",
        "Сбер",
        "Т-Банк",
        "Альфа-Банк",
        "ВТБ",
        "Газпромбанк",
        "Озон Банк",
        "Райффайзен Банк",
        "Lifestyle-конкуренты",
    ]

    def test_unlimited_restaurant_compensation_has_no_fake_period_totals(self):
        evaluation = _compensation_evaluation(
            "Безлимитные компенсации: до 5 000 ₽ на один посадочный талон "
            "в России",
            "restaurants",
        )

        self.assertEqual(evaluation["metrics"]["unlimited"], 1)
        self.assertEqual(evaluation["metrics"]["per_use_limit"], 5000)
        self.assertNotIn("monthly_total", evaluation["metrics"])
        self.assertNotIn("annual_total", evaluation["metrics"])
        self.assertEqual(
            evaluation["summary"],
            "безлимит, до 5 тыс ₽ на один посадочный талон",
        )

    def test_paid_fallback_makes_conditional_service_more_accessible(self):
        flexible = _service_evaluation(
            "Бесплатно при выполнении условий или 2 990 ₽ в месяц"
        )
        conditions_only = _service_evaluation(
            "Бесплатно при выполнении условий"
        )

        self.assertEqual(flexible["metrics"]["service_rank"], 3)
        self.assertEqual(conditions_only["metrics"]["service_rank"], 2)
        self.assertIn("Два способа обслуживания", flexible["summary"])
        self.assertIn("2 990 ₽ в месяц", flexible["summary"])

    def _build_current_outputs(self, tmp):
        workbook = Path(tmp) / "report.xlsx"
        comparison_json = Path(tmp) / "comparison_data.json"
        output = Path(tmp) / "sber_vs.html"
        history = load_history(Path("data/history.json"))
        write_report(history, workbook)
        write_comparison_json(history, comparison_json)
        build_sber_vs_landing(comparison_json, output)
        return workbook, output

    def _user_visible_texts(self, workbook, output):
        wb = load_workbook(workbook, read_only=True, data_only=True)
        for sheet in self.USER_VISIBLE_SHEETS:
            if sheet not in wb.sheetnames:
                continue
            for row in wb[sheet].iter_rows(values_only=True):
                for value in row:
                    if value not in (None, ""):
                        yield f"excel:{sheet}", str(value)

        html = output.read_text(encoding="utf-8")
        payload = re.search(
            r'<script id="data" type="application/json">(.*?)</script>',
            html,
            flags=re.S,
        ).group(1)
        data = self._comparison_payload(payload)
        for bank in data:
            yield "html:bank", bank["bank"]
            for level in bank["levels"]:
                yield f"html:{bank['bank']}", level["tier"]
                for attr in level["attrs"]:
                    yield f"html:{bank['bank']}:{level['tier']}", attr["label"]
                    for key in ("note", "details"):
                        if attr.get(key):
                            yield f"html:{bank['bank']}:{level['tier']}:{attr['id']}:{key}", attr[key]
                    value = attr.get("value")
                    if isinstance(value, list):
                        for item in value:
                            for key in ("title", "description"):
                                if item.get(key):
                                    yield f"html:{bank['bank']}:{level['tier']}:{attr['id']}:{key}", item[key]
                    else:
                        yield f"html:{bank['bank']}:{level['tier']}:{attr['id']}", str(value)

    def _comparison_payload(self, payload: str):
        data = json.loads(payload)
        return data["comparison"] if isinstance(data, dict) else data

    def _write_comparison_json(self, tmp, rows):
        comparison_json = Path(tmp) / "comparison_data.json"
        results = {}
        for idx, row in enumerate(rows, start=1):
            tier_id = row.get("tier_id", f"test_tier_{idx}")
            field_values = row.get("fields", {})
            results[tier_id] = {
                "bank": row["bank"],
                "tier": row["tier"],
                "segment": row.get("segment", "0–3 млн ₽"),
                "scan_date": row.get("scan_date", "2026-07-13T00:00:00"),
                "sources_ok": row.get("sources_ok", 1),
                "source_url": row.get("source_url", "https://example.test"),
                "status": "ok",
                "score": {"total": row.get("score", 0), "breakdown": {}},
                "fields": {fid: merged(value) for fid, value in field_values.items()},
            }
        history = {
            "scans": [{
                "date": "2026-07-13T00:00:00",
                "results": results,
                "meta": {},
            }],
            "changelog": [],
        }
        write_comparison_json(history, comparison_json)
        return comparison_json

    def test_pbi_splits_taxi_restaurants_and_selectable_options(self):
        html = """
        <html><body>
          <h1>Сбер – Новый СберПервый</h1><h3>Уровень за 15 млн ₽</h3>
          <dl>
            <dt>Такси и рестораны</dt>
            <dd>2 в мес по 2000 ₽ — опция «Такси» |
                10 в мес по 4000 ₽ — опция «Рестораны» (всегда включена) |
                До 2 чеков в сутки по 4 тыс ₽</dd>
            <dt>Страховка</dt>
            <dd>$1 млн, 90 дн, ассистанс Mondial Assistance</dd>
            <dt>Привилегии на выбор</dt>
            <dd>Можно выбрать 1 опцию из списка:
                Опция «Авто» (помощь на дорогах, кешбэк 15% за парковки)</dd>
          </dl>
        </body></html>
        """
        fields, quality = parse_source(
            html, {"tier_id": "sber_first_4"}, {"id": "sber", "type": "bank"},
            "pbi", "https://premiumbanking.info/sber/4")

        self.assertEqual(quality, "structured")
        self.assertIn("2 в мес", fields["taxi"])
        self.assertIn("10 в мес", fields["restaurants"])
        self.assertIn("всегда включена", fields["always_included_options"])
        self.assertIn("Опция «Авто»", fields["selectable_options"])
        self.assertIn("выбрать 1 опцию", fields["selection_rules"].lower())
        self.assertIn("$1 млн, 90 дн, ассистанс Mondial Assistance",
                      fields["insurance"])

    def test_pbi_sber_overview_page_extracts_level_specific_blocks(self):
        html = """
        <html><body>
        <h2>Премиальные уровни</h2>
        <h4>Сбер – Уровень 1</h4>
        Новый СберПремьер за 2 млн ₽
        или траты 150 тыс ₽
        или 4000 акций (≈1,180,000₽)
        или 1990 ₽ в мес
        Привилегии на выбор
        Раз в месяц можно выбрать одну из опций пакета.
        Всегда включено:
        —
        Дополнительно на выбор:
        опция «Самокат» — 2 заказа по 500 ₽
        опция «Здоровье» — телемедицина, анализы, исследования
        опция «Авто» — помощь на дорогах, кешбэк 15% за платные дороги и парковки
        Страховка
        Страховки ВЗР в рамках пакета не предусмотрено
        Другие привилегии
        СберПрайм
        <h4>Сбер – Уровень 4</h4>
        Новый СберПервый за 15 млн ₽
        или 40000 акций (≈11,800,000₽)
        или 30000 ₽ в мес
        Привилегии на выбор
        Раз в месяц можно выбрать одну из опций пакета.
        Всегда включено:
        опция «Бизнес-залы» — безлимит
        опция «Рестораны» — 10 ресторанов по 4000 ₽
        Дополнительно на выбор:
        опция «Такси» — 2 такси по 2000 ₽
        Страховка
        $1/1 млн, 90 дн, ассистанс Mondial Assistance
        Другие привилегии
        СберПрайм, Okko «Премиум» с Amediateka
        Консьерж Aspire
        <h4>Сбер – Уровень 6</h4>
        Новый Sber Private за 100 млн ₽
        или 100000 ₽ в мес
        Привилегии на выбор
        Раз в месяц можно выбрать одну из опций пакета.
        Всегда включено:
        опция «Бизнес-залы» — безлимит
        опция «Такси» — 2 такси по 3000 ₽
        опция «Рестораны» — безлимит ресторанов по 5000 ₽
        Дополнительно на выбор:
        опция «Развлечения» — до 8000 ₽ на Афиша.ру
        Страховка
        $1/1 млн, 180 дн, ассистанс Mondial Assistance
        Другие привилегии
        Консьерж Pb Service
        Сбер Мобайл: звонки + 5 гб в месяц
        </body></html>
        """
        bank = {"id": "sber", "name": "Сбер", "type": "bank"}

        level_1, quality = parse_source(
            html, {"tier_id": "sber_premier_1", "tier_name": "СберПремьер — уровень 1"},
            bank, "pbi", "https://premiumbanking.info/sber")
        level_4, _ = parse_source(
            html, {"tier_id": "sber_first_4", "tier_name": "СберПервый — уровень 4"},
            bank, "pbi", "https://premiumbanking.info/sber")
        level_6, _ = parse_source(
            html, {"tier_id": "sber_private_6", "tier_name": "Sber Private Banking — уровень 6"},
            bank, "pbi", "https://premiumbanking.info/sber")

        self.assertEqual(quality, "structured")
        self.assertIn("1990 ₽ в мес", level_1["entry_conditions"])
        self.assertEqual(level_1["service_cost"], "1990 ₽ в мес")
        self.assertIn("Здоровье", level_1["selectable_options"])
        self.assertIn("не предусмотрено", level_1["insurance"])
        self.assertIn("СберПрайм", level_1["ecosystem"])
        self.assertIn("безлимит", level_4["lounge_access"])
        self.assertIn("10 ресторанов по 4000 ₽", level_4["restaurants"])
        self.assertIn("2 такси по 2000 ₽", level_4["taxi"])
        self.assertIn("Pb Service", level_6["concierge"])
        self.assertIn("180 дн", level_6["insurance"])

    def test_availability_comes_from_source_section(self):
        html = """
        <html><body>
        <h4>Сбер – Уровень 4</h4>
        Новый СберПервый за 15 млн ₽
        Всегда включено:
        опция «Бизнес-залы» — безлимит
        опция «Рестораны» — 10 ресторанов по 4000 ₽
        Дополнительно на выбор:
        опция «Такси» — 2 такси по 2000 ₽
        опция «Авто» — помощь на дорогах
        Страховка
        $1 млн, 90 дн
        Другие привилегии
        СберПрайм
        <h4>Сбер – Уровень 5</h4>
        Новый СберПервый за 50 млн ₽
        Всегда включено:
        опция «Бизнес-залы» — безлимит
        опция «Такси» — 2 такси по 3000 ₽
        опция «Рестораны» — 10 ресторанов по 4000 ₽
        Дополнительно на выбор:
        опция «Самокат» — 2 заказа по 2000 ₽
        Страховка
        $1 млн, 90 дн
        Другие привилегии
        СберПрайм
        <h4>Сбер – Уровень 6</h4>
        Новый Sber Private за 100 млн ₽
        Всегда включено:
        опция «Бизнес-залы» — безлимит
        опция «Такси» — 2 такси по 3000 ₽
        опция «Рестораны» — безлимит ресторанов по 5000 ₽
        Дополнительно на выбор:
        опция «Развлечения» — до 8000 ₽
        Страховка
        $1 млн, 180 дн
        Другие привилегии
        СберПрайм
        </body></html>
        """
        bank = {"id": "sber", "name": "Сбер", "type": "bank"}

        for tier_id, tier_name, always_terms, selectable_terms in (
            ("sber_first_4", "СберПервый — уровень 4",
             ("Бизнес-залы", "Рестораны"), ("Такси",)),
            ("sber_first_5", "СберПервый — уровень 5",
             ("Бизнес-залы", "Такси", "Рестораны"), ("Самокат",)),
            ("sber_private_6", "Sber Private Banking — уровень 6",
             ("Бизнес-залы", "Такси", "Рестораны"), ("Развлечения",)),
        ):
            with self.subTest(tier_id=tier_id):
                fields, _ = parse_source(
                    html, {"tier_id": tier_id, "tier_name": tier_name},
                    bank, "pbi", "https://premiumbanking.info/sber")
                for term in always_terms:
                    self.assertIn(term, fields["always_included_options"])
                    self.assertIn("всегда включена", fields["always_included_options"])
                    self.assertNotIn(term, fields["selectable_options"])
                for term in selectable_terms:
                    self.assertIn(term, fields["selectable_options"])

    def test_sber_level_4_availability(self):
        html = """
        <html><body><h4>Сбер – Уровень 4</h4>
        Всегда включено:
        опция «Бизнес-залы» — безлимит
        опция «Рестораны» — 10 ресторанов по 4000 ₽
        Дополнительно на выбор:
        опция «Такси» — 2 такси по 2000 ₽
        Страховка
        $1 млн
        Другие привилегии
        СберПрайм
        </body></html>
        """
        fields, _ = parse_source(
            html, {"tier_id": "sber_first_4", "tier_name": "СберПервый — уровень 4"},
            {"id": "sber", "name": "Сбер", "type": "bank"},
            "pbi", "https://premiumbanking.info/sber")

        self.assertIn("всегда включена", fields["lounge_access"])
        self.assertIn("всегда включена", fields["restaurants"])
        self.assertNotIn("Бизнес-залы", fields["selectable_options"])
        self.assertNotIn("Рестораны", fields["selectable_options"])
        self.assertIn("Такси", fields["selectable_options"])

    def test_sber_level_5_availability(self):
        html = """
        <html><body><h4>Сбер – Уровень 5</h4>
        Всегда включено:
        опция «Бизнес-залы» — безлимит
        опция «Такси» — 2 такси по 3000 ₽
        опция «Рестораны» — 10 ресторанов по 4000 ₽
        Дополнительно на выбор:
        опция «Самокат» — 2 заказа по 2000 ₽
        Страховка
        $1 млн
        Другие привилегии
        СберПрайм
        </body></html>
        """
        fields, _ = parse_source(
            html, {"tier_id": "sber_first_5", "tier_name": "СберПервый — уровень 5"},
            {"id": "sber", "name": "Сбер", "type": "bank"},
            "pbi", "https://premiumbanking.info/sber")

        for field_id in ("lounge_access", "taxi", "restaurants"):
            self.assertIn("всегда включена", fields[field_id])
        self.assertNotIn("Бизнес-залы", fields["selectable_options"])
        self.assertNotIn("Такси", fields["selectable_options"])
        self.assertNotIn("Рестораны", fields["selectable_options"])

    def test_sber_level_6_availability(self):
        html = """
        <html><body><h4>Сбер – Уровень 6</h4>
        Всегда включено:
        опция «Бизнес-залы» — безлимит
        опция «Такси» — 2 такси по 3000 ₽
        опция «Рестораны» — безлимит ресторанов по 5000 ₽
        Дополнительно на выбор:
        опция «Развлечения» — до 8000 ₽
        Страховка
        $1 млн
        Другие привилегии
        СберПрайм
        </body></html>
        """
        fields, _ = parse_source(
            html, {"tier_id": "sber_private_6", "tier_name": "Sber Private Banking — уровень 6"},
            {"id": "sber", "name": "Сбер", "type": "bank"},
            "pbi", "https://premiumbanking.info/sber")

        for field_id in ("lounge_access", "taxi", "restaurants"):
            self.assertIn("всегда включена", fields[field_id])
        self.assertNotIn("Бизнес-залы", fields["selectable_options"])
        self.assertNotIn("Такси", fields["selectable_options"])
        self.assertNotIn("Рестораны", fields["selectable_options"])

    def test_sber_prime_present_all_levels(self):
        blocks = []
        for level in range(1, 7):
            blocks.append(f"""
            <h4>Сбер – Уровень {level}</h4>
            Условия уровня {level}
            Страховка
            $1 млн
            Другие привилегии
            СберПрайм
            """)
        html = "<html><body>" + "\n".join(blocks) + "</body></html>"
        bank = {"id": "sber", "name": "Сбер", "type": "bank"}
        tier_names = {
            1: "СберПремьер — уровень 1",
            2: "СберПремьер — уровень 2",
            3: "СберПремьер — уровень 3",
            4: "СберПервый — уровень 4",
            5: "СберПервый — уровень 5",
            6: "Sber Private Banking — уровень 6",
        }
        tier_ids = {
            1: "sber_premier_1",
            2: "sber_premier_2",
            3: "sber_premier_3",
            4: "sber_first_4",
            5: "sber_first_5",
            6: "sber_private_6",
        }

        for level in range(1, 7):
            fields, _ = parse_source(
                html, {"tier_id": tier_ids[level], "tier_name": tier_names[level]},
                bank, "pbi", "https://premiumbanking.info/sber")
            benefits = build_other_benefits({"ecosystem": merged(fields["ecosystem"])})
            self.assertIn("sber_prime", {item["id"] for item in benefits})

    def test_alfabank_general_page_levels(self):
        html = """
        <html><body>
        <h2>Премиальные уровни</h2>
        <h4>Альфа-Банк – Alfa Only</h4>
        за 3 млн ₽
        или 2 млн ₽ + траты 200 тыс ₽
        Бизнес-залы
        2 в мес (12 в год)
        Рестораны
        2 в мес (12 в год) по 2500 ₽
        Такси
        2 в год (по 2500 ₽)
        Страховка
        $150/35 тыс, 90 дн, ассистанс Class Assistance
        Другие привилегии
        Консьерж-сервис Only Assist
        <h4>Альфа-Банк – А-Клуб</h4>
        за 60 млн ₽ для Мск | 30 млн ₽ регионы
        Бизнес-залы
        безлимит
        Рестораны
        безлимит по 2500 ₽
        Такси
        3 в мес, 15 в год (по 5000 ₽)
        Страховка
        €650 тыс, 90 дн, ассистанс Class Assistance
        Другие привилегии
        Консьерж-сервис PRIME
        </body></html>
        """
        private_level_html = """
        <html><body>
        <dl>
        <dt>Условия</dt>
        <dd>за 50 млн ₽ для Мск | 25 млн ₽ регионы</dd>
        <dt>Бизнес-залы</dt>
        <dd>безлимит</dd>
        <dt>Трансфер</dt>
        <dd>24 в год (по 4000 ₽)</dd>
        <dt>Страховка</dt>
        <dd>$1/1 млн, 365 дн</dd>
        <dt>Другие привилегии</dt>
        <dd>
        <ul>
        <li>Безлимит бизнес-залов</li>
        <li>Трансфер через консьерж-сервис (24 в год по 4000 ₽)</li>
        <li>До 8 бесплатных пакетов private для близких и совместный учет активов</li>
        <li>Ограничено количество гостей в БЗ (до 50М - 1, 50-299М - 3, >300М - 10 гостей)</li>
        <li>Посещение Третьяковской галереи без очереди, без билета, без бронирования</li>
        <li>Консьерж-сервис PRIME</li>
        <li>От 150М: карта World Elite MasterCard Prime</li>
        <li>От 300М: чекап в GMS Clinic или 5 услуг в год сопровождения в аэропорту</li>
        </ul>
        </dd>
        </dl>
        </body></html>
        """
        bank = {"id": "alfa", "name": "Альфа-Банк", "type": "bank"}
        paid, _ = parse_source(
            html, {"tier_id": "alfa_only_1", "tier_name": "Alfa Only — уровень 1"},
            bank, "pbi", "https://premiumbanking.info/alfabank")
        only_3m, _ = parse_source(
            html, {"tier_id": "alfa_only_2", "tier_name": "Alfa Only — уровень 2"},
            bank, "pbi", "https://premiumbanking.info/alfabank")
        aclub, _ = parse_source(
            html, {"tier_id": "alfa_aclub", "tier_name": "A-Club (private)"},
            bank, "pbi", "https://premiumbanking.info/alfabank")

        self.assertEqual(paid["entry_conditions"], NOT_FOUND)
        self.assertIn("3 млн ₽", only_3m["entry_conditions"])
        self.assertIn("Only Assist", only_3m["concierge"])
        self.assertNotIn("PRIME", only_3m["ecosystem"])
        self.assertIn("60 млн ₽", aclub["entry_conditions"])
        self.assertIn("PRIME", aclub["concierge"])

    def test_aclub_structured_fields(self):
        html = """
        <html><body>
        <h4>Альфа-Банк – А-Клуб</h4>
        за 60 млн ₽ для Мск | 30 млн ₽ регионы
        Бизнес-залы
        безлимит
        Рестораны
        безлимит по 2500 ₽
        Такси
        3 в мес, 15 в год (по 5000 ₽)
        Страховка
        €650 тыс, 90 дн, ассистанс Class Assistance
        Другие привилегии
        Консультации с юристом, бухгалтером
        Альфа-Мобайл (50 ГБ, 500 минут, 50 смс)
        Консьерж-сервис PRIME
        Alfa Only Лаундж в SVO терминал С
        А-Клуб Лаундж в SVO терминал B
        </body></html>
        """
        fields, _ = parse_source(
            html, {"tier_id": "alfa_aclub", "tier_name": "A-Club (private)"},
            {"id": "alfa", "name": "Альфа-Банк", "type": "bank"},
            "pbi", "https://premiumbanking.info/alfabank")

        self.assertIn("60 млн ₽", fields["entry_conditions"])
        self.assertIn("безлимит", fields["lounge_access"])
        self.assertIn("безлимит по 2500 ₽", fields["restaurants"])
        self.assertIn("3 в мес, 15 в год", fields["taxi"])
        self.assertIn("€650 тыс", fields["insurance"])
        self.assertIn("Class Assistance", fields["insurance"])
        self.assertIn("PRIME", fields["concierge"])
        self.assertIn("50 ГБ", fields["ecosystem"])
        self.assertIn("А-Клуб Лаундж", fields["ecosystem"])
        self.assertEqual(fields["cashback"], NOT_FOUND)
        self.assertEqual(fields["deposits"], NOT_FOUND)

    def test_alfa_other_benefits_include_services_and_lounges(self):
        benefits = build_other_benefits({
            "ecosystem": merged(
                "Консультации с юристом, бухгалтером ; "
                "Альфа‑Мобайл (10 Гб, 300 минут и 30 смс) ; "
                "Саммари от Smart Reading ; "
                "Подписка РБК ; "
                "Консьерж-сервис Only Assist на платформе Konsierge ; "
                "Alfa Only Лаундж в SVO терминал C"
            ),
        })
        text = json.dumps(benefits, ensure_ascii=False)

        for marker in (
            "Консультации с юристом",
            "Альфа‑Мобайл",
            "Smart Reading",
            "Саммари от Smart Reading",
            "РБК",
            "Подписка",
            "Only Assist",
            "Konsierge",
            "Alfa Only Лаундж",
            "SVO терминал C",
        ):
            self.assertIn(marker, text)
        self.assertNotIn("Консультации с юристом и бухгалтером —", text)

    def test_pbi_rating_reviews_are_not_user_facts(self):
        html = """
        <html><body>
          <dl>
            <dt>Рейтинг уровня по отзывам ПБИ</dt>
            <dd>-</dd>
            <dt>Условия</dt>
            <dd>за 3 млн ₽</dd>
          </dl>
        </body></html>
        """
        fields, _ = parse_source(
            html,
            {"tier_id": "alfa_only_2", "tier_name": "Alfa Only — уровень 2"},
            {"id": "alfa", "name": "Альфа-Банк", "type": "bank"},
            "pbi_level",
            "https://premiumbanking.info/alfabank/2",
        )

        self.assertEqual(fields["other_notes"], NOT_FOUND)
        self.assertEqual(fields["entry_conditions"], "за 3 млн ₽")

    def test_aclub_other_benefits_include_prime_and_aclublounge(self):
        benefits = build_other_benefits({
            "ecosystem": merged(
                "Консультации с юристом, бухгалтером; "
                "Альфа-Мобайл (50 ГБ, 500 минут, 50 смс); "
                "Консьерж-сервис PRIME; "
                "Alfa Only Лаундж в SVO терминал С; "
                "А-Клуб Лаундж в SVO терминал B"
            ),
        })
        text = json.dumps(benefits, ensure_ascii=False)

        for marker in (
            "PRIME",
            "Альфа-Мобайл",
            "50 ГБ",
            "Alfa Only Лаундж",
            "А-Клуб Лаундж",
            "SVO терминал B",
        ):
            self.assertIn(marker, text)

    def test_aclub_other_benefits_include_official_aclub_services(self):
        from scanner.curated import curated_for

        benefits = build_other_benefits({
            "ecosystem": merged(curated_for("alfa_aclub")["ecosystem"]["value"]),
            "concierge": merged(curated_for("alfa_aclub")["concierge"]["value"]),
        })
        text = json.dumps(benefits, ensure_ascii=False)

        for marker in (
            "SimplePrivé",
            "статус Gold",
            "персональный сомелье",
            "скидка 30%",
            "Медицинский консьерж",
            "обследований и лечения",
            "России и за рубежом",
            "PRIME",
        ):
            self.assertIn(marker, text)

    def test_benefit_catalog_does_not_create_missing_benefits(self):
        from scanner.benefit_catalog import classify_benefit

        self.assertIsNotNone(classify_benefit("Консьерж-сервис Only Assist на платформе Konsierge"))
        self.assertIsNotNone(classify_benefit("Закрытый винный клуб SimplePrivé"))
        self.assertIsNotNone(classify_benefit("Медицинский консьерж"))

        benefits = build_other_benefits({
            "ecosystem": merged(NOT_FOUND),
            "always_included_options": merged(NOT_FOUND),
            "selectable_options": merged(NOT_FOUND),
            "auto": merged(NOT_FOUND),
            "concierge": merged(NOT_FOUND),
        })

        self.assertEqual(benefits, [])

    def test_aclub_temporary_promotions_not_used_as_cashback(self):
        fact = curated_for("alfa_aclub")["cashback"]

        self.assertEqual(fact["value"], NOT_FOUND)
        self.assertIn("временные акции не используются", fact["note"])
        self.assertNotIn("ЦУМ", fact["value"])

    def test_aclub_missing_card_terms_not_invented(self):
        facts = curated_for("alfa_aclub")

        self.assertIn("Повышенный доход", facts["deposits"]["value"])
        self.assertIn("alfabank.ru/a-club", facts["deposits"]["source_url"])
        self.assertEqual(facts["card_terms"]["value"], NOT_FOUND)
        self.assertIn("не переносятся", facts["card_terms"]["note"])

    def test_vtb_general_page_levels(self):
        html = """
        <html><body>
        <h4>ВТБ – Привилегия</h4>
        за 2.5 млн ₽ для Мск | 2 млн ₽ регионы
        Привилегии на выбор
        Раз в месяц начисляются «Преференции».
        Преференции
        2 в мес
        Бизнес-залы
        2 в мес, 24 в год — расходует «Преференции»
        Рестораны
        2 в мес (12 в год) по 2500 ₽ — расходует «Преференции»
        Такси
        2 в мес, 12 в год (на 1000 ₽) — расходует «Преференции»
        Страховка
        $100/100 тыс, 90 дн
        Другие привилегии
        Помощь на дорогах (консультации, подвоз топлива, эвакуатор)
        АМА консьерж
        Доступ в сервис appoint
        ·ON·PACK в лимите БЗ
        <h4>ВТБ – Привилегия</h4>
        за 6 млн ₽
        Привилегии на выбор
        Раз в месяц начисляются «Преференции».
        Преференции
        6 в мес
        Бизнес-залы
        6 в мес, 48 в год — расходует «Преференции»
        Рестораны
        6 в мес (24 в год) по 2500 ₽ — расходует «Преференции»
        Такси
        6 в мес, 24 в год (на 1000 ₽) — расходует «Преференции»
        Страховка
        $100/100 тыс, 90 дн
        Другие привилегии
        Помощь на дорогах (консультации, подвоз топлива, эвакуатор)
        АМА консьерж
        5 тыс баллов в год в сервисе appoint
        ·ON·PACK в лимите БЗ
        <h4>ВТБ – Прайм+</h4>
        за 16667 ₽ в мес
        Привилегии на выбор
        Раз в месяц начисляются «Преференции».
        Преференции
        10 в мес
        Бизнес-залы
        10 в мес, 60 в год — расходует «Преференции»
        Рестораны
        10 в мес (30 в год) по 2500 ₽ — расходует «Преференции»
        Такси
        -
        Страховка
        $500 тыс, 90 дн, ассистанс Medlabel / AP Companies
        Другие привилегии
        Помощь на дорогах (консультации, подвоз топлива, эвакуатор)
        АМА консьерж
        Cервис appoint
        Фитмост (120 баллов в год)
        ·ON·PACK в лимите БЗ
        <h4>ВТБ – Прайм+</h4>
        за 100 млн ₽ для Мск | 50 млн ₽ регионы
        Привилегии на выбор
        Раз в месяц начисляются «Преференции».
        Преференции
        25 в мес
        Бизнес-залы
        безлимит
        Рестораны
        25 в мес (50 в год) по 2500 ₽ — расходует «Преференции»
        Такси
        -
        Страховка
        $500 тыс, 90 дн, ассистанс Medlabel / AP Companies
        Другие привилегии
        Помощь на дорогах (консультации, подвоз топлива, эвакуатор)
        АМА консьерж
        Телемедицина «Доктис», онкострахование
        Медицинское обследование 1 раз в год
        Фитмост (120 баллов в год)
        ·ON·PACK
        10 тыс баллов в год в сервисе appoint
        </body></html>
        """
        bank = {"id": "vtb", "name": "ВТБ", "type": "bank"}
        paid_privilege, _ = parse_source(
            html, {"tier_id": "vtb_privilege_1", "tier_name": "Привилегия — уровень 1"},
            bank, "pbi", "https://premiumbanking.info/vtb")
        privilege, _ = parse_source(
            html, {"tier_id": "vtb_privilege_2", "tier_name": "Привилегия — уровень 2"},
            bank, "pbi", "https://premiumbanking.info/vtb")
        prime, _ = parse_source(
            html, {"tier_id": "vtb_prime_5", "tier_name": "Прайм+ — уровень 5"},
            bank, "pbi", "https://premiumbanking.info/vtb")
        prime_top, _ = parse_source(
            html, {"tier_id": "vtb_prime_8", "tier_name": "Прайм+ — уровень 8"},
            bank, "pbi", "https://premiumbanking.info/vtb")

        self.assertEqual(paid_privilege["entry_conditions"], NOT_FOUND)
        self.assertIn("2.5 млн ₽", privilege["entry_conditions"])
        self.assertIn("2 в мес", privilege["selection_rules"])
        self.assertNotIn("10 в мес", privilege["selection_rules"])
        self.assertIn("Помощь на дорогах", privilege["ecosystem"])
        self.assertIn("АМА консьерж", privilege["ecosystem"])
        self.assertIn("Доступ в сервис appoint", privilege["ecosystem"])
        self.assertIn("16667 ₽ в мес", prime["entry_conditions"])
        self.assertIn("10 в мес", prime["restaurants"])
        self.assertEqual(prime["taxi"], NOT_FOUND)
        self.assertIn("Фитмост", prime["ecosystem"])
        self.assertIn("100 млн ₽", prime_top["entry_conditions"])
        self.assertIn("25 в мес", prime_top["selection_rules"])
        self.assertIn("Медицинское обследование 1 раз в год", prime_top["ecosystem"])
        self.assertEqual(prime_top["concierge"], "АМА консьерж")

        benefits = build_other_benefits({"ecosystem": merged(prime_top["ecosystem"])})
        text = json.dumps(benefits, ensure_ascii=False)
        for marker in (
            "Телемедицина",
            "Доктис",
            "Медицинское обследование 1 раз в год",
            "10 тыс баллов в год в сервисе appoint",
        ):
            self.assertIn(marker, text)
        self.assertNotIn("в — год", text)

    def test_gazprombank_general_page_levels(self):
        html = """
        <html><body>
        <h4>Газпромбанк – Премиум</h4>
        за 6 млн ₽
        Привилегии на выбор
        Раз в месяц можно сменить программу.
        Доступно на выбор:
        пакет «Комфортное путешествие» — Бизнес-залы / фаст трек (8 в мес), трансфер Яндекс Go (2 в мес, 8 в год по 2500 ₽), страховка ВЗР
        пакет «Спорт» — WoldClass или X-Fit или DDX Infinity или Фитмост (9000 бонусных рублей в мес на спорт и спа) или ДМС (телемедицина, очные приёмы 4 шт, лабораторная диагностика 1 шт, стационар, скорая помощь), спортивное страхование до 500 000 ₽
        Страховка
        $1/1 млн, 90 дн
        Другие привилегии
        Премиум Консьерж на платформе Konsierge
        <h4>Газпромбанк – Private</h4>
        за 50 млн ₽ для Мск | 25 млн ₽ регионы
        Бизнес-залы
        безлимит
        Трансфер
        24 в год (по 4000 ₽)
        Страховка
        $1/1 млн, 365 дн
        Другие привилегии
        Консьерж-сервис PRIME
        </body></html>
        """
        private_level_html = """
        <html><body>
        <dl>
        <dt>Условия</dt>
        <dd>за 50 млн ₽ для Мск | 25 млн ₽ регионы</dd>
        <dt>Бизнес-залы</dt>
        <dd>безлимит</dd>
        <dt>Трансфер</dt>
        <dd>24 в год (по 4000 ₽)</dd>
        <dt>Страховка</dt>
        <dd>$1/1 млн, 365 дн</dd>
        <dt>Другие привилегии</dt>
        <dd>
        <ul>
        <li>Безлимит бизнес-залов</li>
        <li>Трансфер через консьерж-сервис (24 в год по 4000 ₽)</li>
        <li>До 8 бесплатных пакетов private для близких и совместный учет активов</li>
        <li>Ограничено количество гостей в БЗ (до 50М - 1, 50-299М - 3, >300М - 10 гостей)</li>
        <li>Посещение Третьяковской галереи без очереди, без билета, без бронирования</li>
        <li>Консьерж-сервис PRIME</li>
        <li>От 150М: карта World Elite MasterCard Prime</li>
        <li>От 300М: чекап в GMS Clinic или 5 услуг в год сопровождения в аэропорту</li>
        </ul>
        </dd>
        </dl>
        </body></html>
        """
        bank = {"id": "gazprombank", "name": "Газпромбанк", "type": "bank"}
        premium, _ = parse_source(
            html, {"tier_id": "gpb_premium_2", "tier_name": "Премиум — уровень 2"},
            bank, "pbi", "https://premiumbanking.info/gazprombank")
        private, _ = parse_source(
            html, {"tier_id": "gpb_private", "tier_name": "Private Banking"},
            bank, "pbi", "https://premiumbanking.info/gazprombank")
        private_level, _ = parse_source(
            private_level_html,
            {"tier_id": "gpb_private", "tier_name": "Private Banking"},
            bank, "pbi_level", "https://premiumbanking.info/gazprombank/4")

        self.assertIn("6 млн ₽", premium["entry_conditions"])
        self.assertIn("8 в мес", premium["lounge_access"])
        self.assertIn("трансфер Яндекс Go", premium["taxi"])
        self.assertNotIn("365 дн", premium["insurance"])
        benefits = build_other_benefits({
            "selectable_options": merged(premium["selectable_options"]),
            "ecosystem": merged(premium["ecosystem"]),
        })
        benefits_text = json.dumps(benefits, ensure_ascii=False)
        self.assertNotIn("Пакет «Комфортное путешествие»", benefits_text)
        self.assertIn('"title": "Спорт"', benefits_text)
        self.assertIn("9000 бонусных рублей", benefits_text)
        real_format_benefits = build_other_benefits({
            "ecosystem": merged(
                "Пакет «Комфортное путешествие» (Бизнес-залы / фаст трек "
                "(8 в мес), трансфер Яндекс Go (2 в мес, 8 в год по 2500 ₽), "
                "страховка ВЗР) | Пакет «Спорт» (WoldClass или X-Fit или "
                "DDX Infinity или Фитмост (9000 бонусных рублей в мес на "
                "спорт и спа) или ДМС (телемедицина, очные приёмы 4 шт, "
                "лабораторная диагностика 1 шт, стационар, скорая помощь), "
                "спортивное страхование до 500 000 ₽)"
            )
        })
        real_format_text = json.dumps(real_format_benefits, ensure_ascii=False)
        self.assertNotIn("Пакет «Комфортное путешествие»", real_format_text)
        self.assertIn('"title": "Спорт"', real_format_text)
        self.assertNotIn("трансфер Яндекс Go", real_format_text)
        self.assertIn("365 дн", private["insurance"])
        self.assertIn("24 в год", private["taxi"])
        self.assertIn("Безлимит бизнес-залов", private_level["ecosystem"])
        self.assertIn("Третьяковской галереи", private_level["ecosystem"])
        self.assertIn("World Elite MasterCard Prime", private_level["ecosystem"])
        self.assertIn("GMS Clinic", private_level["ecosystem"])

    def test_gazprombank_travel_option_is_displayed_in_core_rows(self):
        history = {
            "scans": [{
                "date": "2026-07-16",
                "results": {
                    "gpb_premium_2": {
                        "scan_date": "2026-07-16",
                        "fields": {
                            "entry_conditions": merged("6 млн ₽"),
                            "service_cost": merged("0 ₽"),
                            "lounge_access": merged("8 в мес"),
                            "taxi": merged("2 в мес, 8 в год по 2500 ₽"),
                            "insurance": merged("$1/1 млн, 90 дн"),
                            "ecosystem": merged(
                                "Пакет «Комфортное путешествие» (Бизнес-залы / "
                                "фаст трек (8 в мес), трансфер Яндекс Go, "
                                "страховка ВЗР) | Пакет «Спорт» (Фитмост "
                                "9000 бонусных рублей)"
                            ),
                        },
                    }
                },
            }]
        }

        payload = build_comparison_json(history)
        row = next(item for item in payload["rows"] if item["tier_id"] == "gpb_premium_2")

        self.assertTrue(row["fields"]["lounge_access"]["display_value"].startswith(
            "Опция на выбор: Путешествия — "))
        self.assertTrue(row["fields"]["taxi"]["display_value"].startswith(
            "Опция на выбор: Путешествия — "))
        self.assertTrue(row["fields"]["insurance"]["display_value"].startswith(
            "Опция на выбор: Путешествия — "))
        other = row["fields"]["other_benefits"]["display_value"]
        self.assertNotIn("Комфортное путешествие", other)
        self.assertIn("Спорт", other)
        self.assertIn("9000 бонусных рублей", other)

    def test_ozon_general_page_levels(self):
        html = """
        <html><body>
        <h4>Ozon банк – Ultra Bronze</h4>
        за 2 млн ₽
        или 2900 ₽ в мес
        Бизнес-залы
        2 в мес
        Рестораны
        2 в мес по 1000 ₽
        Такси
        -
        Страховка
        €200/50 тыс, 60 дн
        Другие привилегии
        Подписка Ozon Premium
        <h4>Ozon банк – Ultra Platinum</h4>
        за 12 млн ₽
        Бизнес-залы
        безлимит
        Рестораны
        8 в мес по 2500 ₽
        Такси
        -
        Страховка
        €200/50 тыс, 60 дн
        Другие привилегии
        Компенсация Ozon Select (2 в мес на 2500 ₽)
        </body></html>
        """
        bank = {"id": "ozonbank", "name": "Озон Банк", "type": "bank"}
        bronze, _ = parse_source(
            html, {"tier_id": "ozonbank_ultra_bronze", "tier_name": "Ultra Bronze"},
            bank, "pbi", "https://premiumbanking.info/ozon")
        platinum, _ = parse_source(
            html, {"tier_id": "ozonbank_ultra_platinum", "tier_name": "Ultra Platinum"},
            bank, "pbi", "https://premiumbanking.info/ozon")

        self.assertIn("2900 ₽ в мес", bronze["entry_conditions"])
        self.assertIn("2 в мес по 1000 ₽", bronze["restaurants"])
        self.assertEqual(bronze["taxi"], NOT_FOUND)
        self.assertIn("12 млн ₽", platinum["entry_conditions"])
        self.assertIn("8 в мес", platinum["restaurants"])

    def test_raiffeisen_general_page_levels(self):
        html = """
        <html><body>
        <h4>Райффайзен – Premium</h4>
        за 2500 ₽ в мес
        Бизнес-залы
        4 в год
        Рестораны
        -
        Такси
        -
        Страховка
        $150/150 тыс, 30 дн
        Другие привилегии
        Автоконсьерж
        <h4>Райффайзен – Premium</h4>
        за 5 млн ₽
        Бизнес-залы
        безлимит
        Рестораны
        20 в мес (20 в год) по 2500 ₽
        Такси
        -
        Страховка
        $150/150 тыс, 30 дн
        Другие привилегии
        Сертификаты на 7000
        </body></html>
        """
        bank = {"id": "raiffeisen", "name": "Райффайзен Банк", "type": "bank"}
        paid, _ = parse_source(
            html, {"tier_id": "raif_premium_1", "tier_name": "Premium (за плату)"},
            bank, "pbi", "https://premiumbanking.info/raiffeisen")
        balance, _ = parse_source(
            html, {"tier_id": "raif_premium_4", "tier_name": "Premium (за 5 млн ₽)"},
            bank, "pbi", "https://premiumbanking.info/raiffeisen")

        self.assertIn("2500 ₽ в мес", paid["entry_conditions"])
        self.assertIn("4 в год", paid["lounge_access"])
        self.assertNotIn("20 в мес", paid["restaurants"])
        self.assertIn("5 млн ₽", balance["entry_conditions"])
        self.assertIn("20 в мес", balance["restaurants"])

    def test_tbank_general_page_levels(self):
        html = """
        <html><body>
        <h4>Т-Банк – Diamond</h4>
        Уровень за 10 млн ₽
        Бизнес-залы
        безлимит
        Рестораны
        10 в мес по 2000 ₽
        Такси
        2 в мес (по 2500 ₽)
        Страховка
        $100/100 тыс, 45 дн
        Другие привилегии
        1 уровень Bronze для друга
        <h4>Т-Банк – Private</h4>
        Уровень за 55 млн ₽
        Бизнес-залы
        безлимит
        Рестораны
        безлимит по 5000 ₽
        Такси
        3 в мес (по 8000 ₽)
        Страховка
        $200/200 тыс, 90 дн
        Другие привилегии
        3 уровня Bronze для друзей
        </body></html>
        """
        bank = {"id": "tbank", "name": "Т-Банк", "type": "bank"}
        diamond, _ = parse_source(
            html, {"tier_id": "tbank_diamond", "tier_name": "Premium Diamond"},
            bank, "pbi", "https://premiumbanking.info/tbank")

        self.assertIn("10 млн ₽", diamond["entry_conditions"])
        self.assertIn("2 в мес", diamond["taxi"])
        self.assertNotIn("8000 ₽", diamond["taxi"])
        self.assertNotIn("$200/200", diamond["insurance"])

    def test_tbank_bronze_level_page_does_not_publish_balance_accounting_label(self):
        html = """
        <html><body>
        <h1>Т-Банк – Bronze</h1>
        <h3>Уровень за 2990 ₽ в мес или 5000 акций (≈1,220,000₽)</h3>
        <dl>
          <dt>Условия обслуживания</dt>
          <dd><b>2990 ₽ в мес</b><p>или 5000 акций (≈1,220,000₽)</p></dd>
          <dt>Учет остатков</dt>
          <dd>Минимальный остаток</dd>
        </dl>
        </body></html>
        """
        bank = {"id": "tbank", "name": "Т-Банк", "type": "bank"}
        bronze, _ = parse_source(
            html, {"tier_id": "tbank_bronze", "tier_name": "Premium Bronze"},
            bank, "pbi_level", "https://premiumbanking.info/tbank/1")

        self.assertIn("2990 ₽ в мес", bronze["entry_conditions"])
        self.assertIn("5000 акций", bronze["entry_conditions"])
        self.assertNotIn("Минимальный остаток", bronze["entry_conditions"])

    def test_tbank_bronze_entry_conditions_keep_fee_and_shares_in_landing(self):
        summary = _condition_summary(
            "2990 ₽ в мес; или 5000 акций (≈1,220,000₽)",
            "tbank_bronze",
        )

        self.assertEqual(
            summary,
            "Уровень за 2990 ₽ в мес\nили 5000 акций (≈1,220,000₽)",
        )

    def test_raiffeisen_paid_entry_conditions_show_monthly_fee_only(self):
        summary = _condition_summary(
            "2500 ₽ в мес; Весь последний календарный день месяца",
            "raif_premium_1",
        )

        self.assertEqual(summary, "2 500 ₽ в месяц")

    def test_raiffeisen_paid_service_cost_show_monthly_fee_only(self):
        summary = _service_cost_summary({
            "tier_id": "raif_premium_1",
            "fields": {
                "entry_conditions": "2500 ₽ в мес; Весь последний календарный день месяца",
                "service_cost": "Бесплатно при выполнении условий или 2500 ₽ в мес",
            },
        })

        self.assertEqual(summary["display"], "2 500 ₽ в месяц")

    def test_general_page_is_primary_pbi_source_all_banks(self):
        expected = {
            "alfa": PRIORITY_SOURCE_URLS["pbi"]["alfabank"],
            "vtb": PRIORITY_SOURCE_URLS["pbi"]["vtb"],
            "gazprombank": PRIORITY_SOURCE_URLS["pbi"]["gazprombank"],
            "ozonbank": PRIORITY_SOURCE_URLS["pbi"]["ozon"],
            "raiffeisen": PRIORITY_SOURCE_URLS["pbi"]["raiffeisen"],
            "tbank": PRIORITY_SOURCE_URLS["pbi"]["tbank"],
        }
        banks = {bank["id"]: bank for bank in BANKS}
        for bank_id, overview_url in expected.items():
            with self.subTest(bank_id=bank_id):
                for tier in banks[bank_id]["tiers"]:
                    pbi_sources = [s for s in tier["sources"] if s["source_id"] == "pbi"]
                    self.assertEqual(pbi_sources[0]["urls"][0], overview_url)

    def test_level_pages_are_preferred_over_general_pbi(self):
        banks = {bank["id"]: bank for bank in BANKS}
        checked = 0
        for bank_id in ("alfa", "vtb", "gazprombank", "ozonbank", "raiffeisen", "tbank"):
            for tier in banks[bank_id]["tiers"]:
                pbi = next(s for s in tier["sources"] if s["source_id"] == "pbi")
                self.assertGreaterEqual(len(pbi["urls"]), 2)
                self.assertRegex(pbi["urls"][1], r"/\d+$")
                runtime_pbi = [
                    s for s in tier_sources(banks[bank_id], tier)
                    if s["source_id"] in {"pbi", "pbi_level"}
                ]
                self.assertEqual(runtime_pbi[0]["source_id"], "pbi_level")
                self.assertEqual(runtime_pbi[1]["source_id"], "pbi")
                self.assertEqual(runtime_pbi[0]["urls"], [pbi["urls"][1]])
                self.assertEqual(runtime_pbi[1]["urls"], [pbi["urls"][0]])
                checked += 1
        self.assertGreater(checked, 0)

    def test_no_cross_product_leak(self):
        html = """
        <html><body>
        <h4>Альфа-Банк – Alfa Only</h4>
        за 3 млн ₽
        Страховка
        only insurance
        Другие привилегии
        Only Assist
        <h4>Альфа-Банк – А-Клуб</h4>
        за 60 млн ₽
        Страховка
        club insurance
        Другие привилегии
        PRIME
        <h4>ВТБ – Привилегия</h4>
        за 2.5 млн ₽
        Страховка
        privilege insurance
        <h4>ВТБ – Прайм+</h4>
        за 16667 ₽ в мес
        Страховка
        prime insurance
        <h4>Т-Банк – Diamond</h4>
        Уровень за 10 млн ₽
        Страховка
        diamond insurance
        <h4>Т-Банк – Private</h4>
        Уровень за 55 млн ₽
        Страховка
        private insurance
        </body></html>
        """
        alfa, _ = parse_source(
            html, {"tier_id": "alfa_only_2", "tier_name": "Alfa Only — уровень 2"},
            {"id": "alfa", "name": "Альфа-Банк", "type": "bank"},
            "pbi", "https://premiumbanking.info/alfabank")
        vtb, _ = parse_source(
            html, {"tier_id": "vtb_privilege_2", "tier_name": "Привилегия — уровень 2"},
            {"id": "vtb", "name": "ВТБ", "type": "bank"},
            "pbi", "https://premiumbanking.info/vtb")
        tbank, _ = parse_source(
            html, {"tier_id": "tbank_diamond", "tier_name": "Premium Diamond"},
            {"id": "tbank", "name": "Т-Банк", "type": "bank"},
            "pbi", "https://premiumbanking.info/tbank")

        self.assertEqual(alfa["insurance"], "only insurance")
        self.assertEqual(vtb["insurance"], "privilege insurance")
        self.assertEqual(tbank["insurance"], "diamond insurance")

    def test_all_general_pages_registered(self):
        urls = {
            PRIORITY_SOURCE_URLS["pbi"]["alfabank"],
            PRIORITY_SOURCE_URLS["pbi"]["vtb"],
            PRIORITY_SOURCE_URLS["pbi"]["gazprombank"],
            PRIORITY_SOURCE_URLS["pbi"]["ozon"],
            PRIORITY_SOURCE_URLS["pbi"]["raiffeisen"],
            PRIORITY_SOURCE_URLS["pbi"]["tbank"],
        }
        configured = {
            source["urls"][0]
            for bank in BANKS
            for tier in bank["tiers"]
            for source in tier["sources"]
            if source["source_id"] == "pbi"
        }
        self.assertTrue(urls <= configured)

    def test_score_does_not_sum_taxi_and_restaurant_counts(self):
        fields = {
            "lounge_access": merged(NOT_FOUND),
            "cashback": merged(NOT_FOUND),
            "deposits": merged(NOT_FOUND),
            "taxi": merged("2 в мес по 2000 ₽ — опция «Такси»"),
            "restaurants": merged("10 в мес по 4000 ₽ — всегда включена"),
            "insurance": merged(NOT_FOUND),
            "concierge": merged(NOT_FOUND),
            "ecosystem": merged(NOT_FOUND),
            "auto": merged("Опция «Авто» (помощь на дорогах)"),
        }
        breakdown = score_tier(fields)["breakdown"]

        metric = breakdown["taxi_restaurants"]["metric"]
        self.assertIn("10 в месяц", metric)
        self.assertNotIn("12", metric)
        self.assertEqual(breakdown["auto"]["score"], 1)

    def test_html_uses_other_benefits_list_and_keeps_exact_insurance(self):
        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "sber_vs.html"
            comparison_json = self._write_comparison_json(tmp, [
                {
                    "tier_id": "sber_first_4",
                    "bank": "Сбер",
                    "tier": "СберПервый — уровень 4",
                    "segment": "10–25 млн ₽",
                    "score": 4.7,
                    "fields": {
                        "entry_conditions": "15 млн ₽",
                        "service_cost": "30000 ₽ в мес",
                        "lounge_access": "безлимит",
                        "cashback": "до 10%",
                        "deposits": "до 13%",
                        "taxi": "2 в мес по 2000 ₽ — опция «Такси»",
                        "restaurants": "10 в мес по 4000 ₽ — всегда включена",
                        "insurance": "$1 млн, 90 дн, ассистанс Mondial Assistance",
                        "concierge": "Aspire",
                        "other_benefits": (
                            "• СберПрайм\n• Авто — помощь [опция на выбор]\n"
                            "Условия выбора: одна опция в месяц"
                        ),
                    },
                },
                {
                    "tier_id": "alfa_only_1",
                    "bank": "Альфа-Банк",
                    "tier": "Alfa Only — уровень 1",
                    "score": 2.1,
                    "fields": {
                        "entry_conditions": "2990 ₽ в мес",
                        "service_cost": "2990 ₽ в мес",
                        "lounge_access": "-",
                        "cashback": NOT_FOUND,
                        "deposits": NOT_FOUND,
                        "taxi": "-",
                        "restaurants": "2 в мес по 2500 ₽",
                        "insurance": "$150/35 тыс, 90 дн, ассистанс Class Assistance",
                        "concierge": "Есть",
                        "other_benefits": "• Альфа-Мобайл — 10 ГБ",
                    },
                },
            ])

            build_sber_vs_landing(comparison_json, output)
            html = output.read_text(encoding="utf-8")

        self.assertNotIn('<p class="seg"', html)
        self.assertIn('"label": "Такси"', html)
        self.assertIn('"label": "Рестораны"', html)
        self.assertIn('"label": "Другие привилегии"', html)
        self.assertNotIn('"label": "Опции на выбор"', html)
        self.assertNotIn('"label": "Правила выбора"', html)
        self.assertNotIn('"label": "Авто"', html)
        self.assertIn('benefits-list', html)
        self.assertIn("$150/35 тыс, 90 дн, ассистанс Class Assistance", html)
        match = re.search(
            r'<script id="data" type="application/json">(.*?)</script>',
            html,
            flags=re.S,
        )
        self.assertIsNotNone(match)
        data = self._comparison_payload(match.group(1))
        attrs = data[0]["levels"][0]["attrs"]
        other = next(a for a in attrs if a["label"] == "Другие привилегии")
        self.assertEqual(other["kind"], "benefits")
        self.assertEqual(other["value"][1]["availability"], "selectable")

    def test_landing_uses_json_as_single_source(self):
        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "sber_vs.html"
            comparison_json = self._write_comparison_json(tmp, [
                {
                    "tier_id": "sber_premier_1",
                    "bank": "Сбер",
                    "tier": "СберПремьер — уровень 1",
                    "fields": {
                        "entry_conditions": "1,5 млн ₽",
                        "service_cost": "0 ₽",
                        "lounge_access": NOT_FOUND,
                        "cashback": NOT_FOUND,
                        "deposits": NOT_FOUND,
                        "taxi": NOT_FOUND,
                        "restaurants": NOT_FOUND,
                        "insurance": "$1 млн, визовые риски, потеря багажа",
                        "concierge": "Aspire",
                        "other_benefits": NOT_FOUND,
                    },
                },
                {
                    "tier_id": "tbank_premium_bronze",
                    "bank": "Т-Банк",
                    "tier": "Premium Bronze",
                    "fields": {
                        "entry_conditions": "1 млн ₽",
                        "service_cost": "0 ₽",
                        "lounge_access": NOT_FOUND,
                        "cashback": NOT_FOUND,
                        "deposits": NOT_FOUND,
                        "taxi": NOT_FOUND,
                        "restaurants": NOT_FOUND,
                        "insurance": NOT_FOUND,
                        "concierge": NOT_FOUND,
                        "other_benefits": "• T Premium",
                    },
                },
            ])

            build_sber_vs_landing(comparison_json, output)
            html = output.read_text(encoding="utf-8")

        match = re.search(
            r'<script id="data" type="application/json">(.*?)</script>',
            html,
            flags=re.S,
        )
        self.assertIsNotNone(match)
        data = self._comparison_payload(match.group(1))
        sber_attrs = data[0]["levels"][0]["attrs"]
        other = next(a for a in sber_attrs if a["label"] == "Другие привилегии")
        self.assertEqual(other["kind"], "benefits")
        self.assertEqual(other["value"], [])
        self.assertNotIn("Aspire", json.dumps(other, ensure_ascii=False))

    def test_new_bank_fields_flow_to_json_excel_and_html_with_provenance(self):
        with tempfile.TemporaryDirectory() as tmp:
            workbook = Path(tmp) / "report.xlsx"
            comparison_json = Path(tmp) / "comparison_data.json"
            output = Path(tmp) / "sber_vs.html"
            history = {
                "scans": [{
                    "date": "2026-07-17T00:00:00",
                    "results": {
                        "sber_premier_1": {
                            "bank": "Сбер",
                            "tier": "СберПремьер — уровень 1",
                            "segment": "0–3 млн ₽",
                            "scan_date": "2026-07-17T00:00:00",
                            "sources_ok": 1,
                            "source_url": "https://bank.example/sber",
                            "status": "ok",
                            "score": {"total": 0, "breakdown": {}},
                            "fields": {
                                **self._comparison_fields(),
                                "transfers_payments": {
                                    "value": "переводы без комиссии до 1 млн ₽",
                                    "source_id": "official",
                                    "source_type": "official",
                                    "source_url": "https://bank.example/sber/transfers",
                                    "date_checked": "2026-07-17",
                                    "raw_text": "переводы без комиссии до 1 млн ₽",
                                },
                                "cash_withdrawal": {
                                    "value": "снятие наличных без комиссии до 500 000 ₽",
                                    "source_id": "official",
                                    "source_type": "official",
                                    "source_url": "https://bank.example/sber/cash",
                                    "date_checked": "2026-07-17",
                                    "raw_text": "снятие наличных без комиссии до 500 000 ₽",
                                },
                                "supreme": {
                                    "value": "Карта Мир Supreme включена",
                                    "source_id": "official",
                                    "source_type": "official",
                                    "source_url": "https://bank.example/sber/supreme",
                                    "date_checked": "2026-07-17",
                                    "raw_text": "Карта Мир Supreme включена",
                                },
                            },
                        },
                        "alfa_only_1": {
                            "bank": "Альфа-Банк",
                            "tier": "Alfa Only — уровень 1",
                            "segment": "0–3 млн ₽",
                            "scan_date": "2026-07-17T00:00:00",
                            "sources_ok": 1,
                            "source_url": "https://bank.example/alfa",
                            "status": "ok",
                            "score": {"total": 0, "breakdown": {}},
                            "fields": self._comparison_fields(
                                transfers_payments=NOT_FOUND,
                                cash_withdrawal=NOT_FOUND,
                                supreme=NOT_FOUND,
                            ),
                        },
                    },
                    "meta": {},
                }],
                "changelog": [],
            }

            write_report(history, workbook)
            payload = write_comparison_json(history, comparison_json)
            with patch("landing.sber_vs.premium_changes.load_changes", return_value=[]):
                build_sber_vs_landing(comparison_json, output)
            wb = load_workbook(workbook, read_only=True, data_only=True)
            html = output.read_text(encoding="utf-8")

        row = next(item for item in payload["rows"] if item["tier_id"] == "sber_premier_1")
        for field_id in ("transfers_payments", "cash_withdrawal", "supreme"):
            self.assertIn(field_id, row["fields"])
            self.assertEqual(row["fields"][field_id]["source_type"], "official")
            self.assertEqual(row["fields"][field_id]["date_checked"], "2026-07-17")
            self.assertIn("https://bank.example/sber/", row["fields"][field_id]["source_url"])
            self.assertTrue(row["fields"][field_id]["raw_text"])
            self.assertIn(row["fields"][field_id]["display_value"], html)

        summary_header = next(wb["Сводная"].iter_rows(min_row=1, max_row=1, values_only=True))
        self.assertIn("Переводы и платежи без комиссии", summary_header)
        self.assertIn("Снятие наличных", summary_header)
        self.assertIn("Supreme", summary_header)
        self.assertIn("Не найдено в доступных источниках", html)
        self.assertIn(
            "const ALWAYS_SHOW_FIELDS = new Set(['transfers_payments', 'cash_withdrawal', 'supreme']);",
            html,
        )
        self.assertIn("!ALWAYS_SHOW_FIELDS.has(baseAttr.id)", html)

    def test_pbi_parser_extracts_new_bank_fields_without_cross_tier_transfer(self):
        html = """
        <html><body>
          <h1>ВТБ – Привилегия</h1><h3>Уровень за 2.5 млн ₽</h3>
          <dl>
            <dt>Переводы и платежи</dt>
            <dd>Переводы без комиссии до 2 млн ₽ в месяц</dd>
            <dt>Снятие наличных</dt>
            <dd>Снятие наличных без комиссии до 1 млн ₽ в месяц</dd>
            <dt>Supreme</dt>
            <dd>Карта Привилегия Mir Supreme</dd>
          </dl>
        </body></html>
        """
        tier = {"tier_id": "vtb_privilege_2", "tier_name": "Привилегия — уровень 2"}
        bank = {"id": "vtb", "name": "ВТБ", "type": "bank"}

        parsed, parse_type = parse_source(
            html, tier, bank, "pbi", "https://premiumbanking.info/vtb/2")

        self.assertEqual(parse_type, "structured")
        self.assertIn("2 млн ₽", parsed["transfers_payments"])
        self.assertIn("1 млн ₽", parsed["cash_withdrawal"])
        self.assertIn("Mir Supreme", parsed["supreme"])
        self.assertNotEqual(parsed["supreme"], NOT_FOUND)

    def test_sber_level_1_other_benefits_exact_set(self):
        benefits = build_other_benefits({
            "always_included_options": merged("СберПрайм"),
            "selectable_options": merged(
                "Включено исследование до 1 тыс. | "
                "Опция «Самокат» (2 заказа по 500 ₽) | "
                "Опция «Здоровье» (телемедицина, анализы, исследования) | "
                "Опция «Питомцы» (лечение и консультации) | "
                "Опция «Авто» (помощь на дорогах, кэшбэк 15% за платные дороги и парковки)"
            ),
            "selection_rules": merged("Раз в месяц можно выбрать одну из опций пакета"),
            "insurance": merged(
                "страховая сумма, визовые риски, задержка или отмена рейса, "
                "потеря багажа, горные лыжи и сноуборд"
            ),
        })

        self.assertEqual(
            {item["id"] for item in benefits},
            {"sber_prime", "samokat", "health", "pets", "auto"},
        )

    def test_sber_level_1_no_insurance_leak(self):
        benefits = build_other_benefits({
            "selectable_options": merged(
                "Опция «Самокат» (2 заказа по 500 ₽) | "
                "Опция «Визовые риски» (страхование визовых рисков) | "
                "Опция «Задержка рейса» (задержка или отмена рейса) | "
                "Опция «Багаж» (потеря багажа) | "
                "Опция «Горные лыжи» (катание на горных лыжах и сноуборде) | "
                "Опция «Assistance» (ассистанс и страховая сумма)"
            ),
        })

        dumped = json.dumps(benefits, ensure_ascii=False).lower()
        for marker in (
            "visa", "flight_delay", "baggage", "ski", "snowboard",
            "insurance", "assistance", "визов", "рейс", "багаж",
            "горные лыжи", "сноуборд", "страхован", "ассистанс",
        ):
            self.assertNotIn(marker, dumped)
        self.assertEqual({item["id"] for item in benefits}, {"samokat"})

    def test_sber_levels_4_5_6_curated_other_benefits(self):
        from scanner.curated import curated_for

        expected = {
            "sber_first_4": ("СберПрайм", "Okko", "Бизнес-зал Сбер в SVO", "Компенсация БЗ"),
            "sber_first_5": ("СберПрайм", "Okko", "Бизнес-зал Сбер в SVO", "Компенсация БЗ"),
            "sber_private_6": ("СберПрайм", "Okko", "СберПраво", "Сбер Мобайл", "Pb Service"),
        }

        for tier_id, markers in expected.items():
            fields = {
                "ecosystem": merged(curated_for(tier_id)["ecosystem"]["value"]),
                "selection_rules": merged("Раз в месяц можно выбрать одну из опций пакета"),
            }
            text = json.dumps(build_other_benefits(fields), ensure_ascii=False)
            with self.subTest(tier_id=tier_id):
                for marker in markers:
                    self.assertIn(marker, text)
                self.assertNotIn("Ð", text)

    def test_alfa_pdf_source_and_cashback_curated(self):
        alfa = next(bank for bank in BANKS if bank["id"] == "alfa")
        urls = [
            url
            for tier in alfa["tiers"][:4]
            for source in tier["sources"]
            for url in source["urls"]
        ]
        self.assertTrue(any("Tariffs_Alfa_Only_Card.pdf" in url for url in urls))

        fact = curated_for("alfa_only_2")["cashback"]
        self.assertIn("7% в 5 категориях", fact["value"])
        self.assertIn("7% в 4 категориях + 1% на всё", fact["value"])
        self.assertIn("30 000 ₽", fact["value"])
        self.assertIn("суперкэшбэк до 100%", fact["value"])
        deposits = curated_for("alfa_only_2")["deposits"]
        self.assertIn("Премиальный вклад", deposits["value"])
        self.assertIn("alfabank.ru/everyday/alfa-only", deposits["source_url"])

        alfa_only = curated_for("alfa_only_2")
        self.assertIn("до 100 000 ₽ в месяц", alfa_only["transfers_payments"]["value"])
        self.assertIn("Снятие наличных", alfa_only["cash_withdrawal"]["value"])
        self.assertIn("Supreme в тарифе карты Alfa Only не заявлен", alfa_only["supreme"]["value"])
        self.assertIn("Tariffs_Alfa_Only_Card.pdf", alfa_only["supreme"]["source_url"])

    def test_tbank_pdf_transfer_cash_and_supreme_curated(self):
        premium = curated_for("tbank_gold")
        private = curated_for("tbank_private_30")

        self.assertIn("100 000 ₽", premium["transfers_payments"]["value"])
        self.assertIn("200 000 ₽", premium["transfers_payments"]["value"])
        self.assertIn("в 2 раза больше", premium["cash_withdrawal"]["value"])
        self.assertIn("активном сервисе Premium", premium["supreme"]["value"])
        self.assertIn("docs-terms-of-service-premium.pdf", premium["supreme"]["source_url"])

        self.assertIn("500 000 ₽", private["transfers_payments"]["value"])
        self.assertIn("в 10 раз больше", private["cash_withdrawal"]["value"])
        self.assertIn("активном сервисе Private", private["supreme"]["value"])
        self.assertIn("docs-terms-of-service-private.pdf", private["supreme"]["source_url"])

    def test_ozon_ultra_pdf_transfer_cash_and_supreme_curated(self):
        bronze = curated_for("ozonbank_ultra_bronze")
        platinum = curated_for("ozonbank_ultra_platinum")

        self.assertIn("MCC 6538", bronze["transfers_payments"]["value"])
        self.assertIn("отдельный лимит", bronze["transfers_payments"]["value"])
        self.assertIn("3 000 000 ₽", bronze["cash_withdrawal"]["value"])
        self.assertIn("Mir Supreme", bronze["supreme"]["value"])
        self.assertIn("Тариф%20Ultra.pdf", bronze["supreme"]["source_url"])

        self.assertIn("30 000 000 ₽", platinum["cash_withdrawal"]["value"])
        self.assertIn("металлическая карта Mir Supreme", platinum["supreme"]["value"])

    def test_summary_never_cuts_word(self):
        text = ("Кэшбэк по карте Alfa Only: 7% в 5 категориях либо 7% в 4 "
                "категориях + 1% на всё; максимальный месячный лимит "
                "кэшбэка 30 000 ₽; суперкэшбэк до 100% как отдельная "
                "промо-механика для партнёров.")
        summary = make_complete_summary(text, 115)

        self.assertNotRegex(summary, r"[А-Яа-яA-Za-z]-?ме[.…]$")
        self.assertNotRegex(summary, r"\w…$")
        self.assertFalse(summary.endswith((" и.", " в.", " на.", " как.")))

    def test_summary_ends_on_complete_phrase(self):
        text = ("Первое законченное условие; второе законченное условие с "
                "важным лимитом 30 000 ₽; третье условие слишком длинное для "
                "краткого отображения и должно уйти в детали.")
        summary = make_complete_summary(text, 82)

        self.assertRegex(summary, r"[.;!?]$")
        self.assertNotIn("слишком длин", summary)

    def test_format_service_list(self):
        self.assertEqual(
            format_list(["Mir Pass", "ON·PASS", "Phoenix Pass", "Soft Travel",
                         "Only Assist"]),
            "Mir Pass, ON·PASS, Phoenix Pass, Soft Travel и Only Assist",
        )
        self.assertEqual(
            format_natural_list(["Mir Pass", "ON·PASS"]),
            "Mir Pass и ON·PASS",
        )

    def test_cleanup_raw_access_text(self):
        raw = ("2 в мес (12 в год), доступ через | Mir Pass | ( | ·ON·PASS | ), "
               "| Phoenix Pass | , | Soft travel | , Only Assist")
        formatted = normalize_source_text(raw)

        self.assertIn("2 посещения в месяц, до 12 в год.", formatted)
        self.assertIn("Доступ через Mir Pass, ON·PASS, Phoenix Pass, Soft Travel и Only Assist.", formatted)
        self.assertNotIn("|", formatted)
        self.assertNotIn("( , )", formatted)

    def test_question_references_are_hidden_from_user_text(self):
        raw = ("Компенсация оплаты за последние 30 дней по кнопке в приложении, "
               "подробнее; в вопросе #36; .")
        formatted = normalize_source_text(raw)

        self.assertNotIn("в вопросе", formatted.lower())
        self.assertNotIn("подробнее", formatted.lower())
        self.assertEqual(
            formatted,
            "Компенсация оплаты за последние 30 дней по кнопке в приложении.",
        )
        self.assertEqual(
            normalize_source_text("Дневного лимита компенсаций нет, подробнее."),
            "Дневного лимита компенсаций нет.",
        )

    def test_cashback_summary_is_complete(self):
        raw = ("Кэшбэк по карте Alfa Only: 7% в 5 категориях либо 7% в 4 "
               "категориях + 1% на всё; максимальный месячный лимит "
               "кэшбэка 30 000 ₽; суперкэшбэк до 100% как отдельная "
               "промо-механика для партнёров.")
        summary = make_complete_summary(raw, 175)

        self.assertNotIn("промо-ме", summary)
        self.assertRegex(summary, r"[.;!?]$")

    def test_full_details_preserved(self):
        raw = ("Первая часть; вторая часть; полный текст с лимитом 30 000 ₽ "
               "и дополнительным условием для отдельных предложений.")
        split = split_summary_and_details(raw, 35)

        self.assertIn("30 000 ₽", split["details"])
        self.assertIn("дополнительным условием", split["details"])
        self.assertNotEqual(split["summary"], split["details"])

    def test_html_has_expandable_details_for_long_text(self):
        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "sber_vs.html"
            long_text = ("Кэшбэк по карте Alfa Only: 7% в 5 категориях либо "
                         "7% в 4 категориях + 1% на всё; максимальный "
                         "месячный лимит кэшбэка 30 000 ₽; суперкэшбэк до "
                         "100% как отдельная промо-механика для партнёров.")
            comparison_json = self._write_comparison_json(tmp, [
                {
                    "tier_id": "alfa_only_1",
                    "bank": "Альфа-Банк",
                    "tier": "Alfa Only — уровень 1",
                    "score": 2.1,
                    "fields": {
                        "entry_conditions": "2990 ₽ в мес",
                        "service_cost": "2990 ₽ в мес",
                        "lounge_access": "2 в мес (12 в год), доступ через | Mir Pass | ( | ·ON·PASS | ), | Phoenix Pass | , | Soft travel | , Only Assist",
                        "cashback": long_text,
                        "deposits": NOT_FOUND,
                        "taxi": NOT_FOUND,
                        "restaurants": NOT_FOUND,
                        "insurance": NOT_FOUND,
                        "concierge": NOT_FOUND,
                        "other_benefits": NOT_FOUND,
                    },
                },
                {
                    "tier_id": "sber_premier_1",
                    "bank": "Сбер",
                    "tier": "СберПремьер — уровень 1",
                    "score": 1.2,
                    "fields": {
                        "entry_conditions": "1,5 млн ₽",
                        "service_cost": "0 ₽",
                        "lounge_access": NOT_FOUND,
                        "cashback": NOT_FOUND,
                        "deposits": NOT_FOUND,
                        "taxi": NOT_FOUND,
                        "restaurants": NOT_FOUND,
                        "insurance": NOT_FOUND,
                        "concierge": NOT_FOUND,
                        "other_benefits": NOT_FOUND,
                    },
                },
            ])

            build_sber_vs_landing(comparison_json, output)
            html = output.read_text(encoding="utf-8")

        self.assertIn("function appendDetails", html)
        self.assertIn("Подробнее", html)
        self.assertIn("attr-details", html)
        self.assertIn('"details":', html)
        self.assertNotIn("| Mir Pass |", html)

    def test_no_css_only_truncation(self):
        html = Path("landing/sber_vs.py").read_text(encoding="utf-8")

        self.assertNotIn("line-clamp", html)
        self.assertNotIn("text-overflow", html)

    def test_pdf_export_uses_one_a3_page_for_full_comparison(self):
        source = Path("landing/sber_vs.py").read_text(encoding="utf-8")

        self.assertIn("function pdfCaptureSizeForElement", source)
        self.assertIn("maxCanvasSide = 16000", source)
        self.assertIn("format: 'a3'", source)
        self.assertIn("orientation: 'landscape'", source)
        self.assertIn("const imageY = marginMm", source)
        self.assertIn("pdf.addImage(", source)
        self.assertIn("pdf.save(comparePdfFileName())", source)
        self.assertNotIn("await html2pdf()", source)
        self.assertNotIn("format: 'a4'", source)
        self.assertNotIn("pagebreak:", source)

    def test_no_pipe_in_user_visible_text(self):
        with tempfile.TemporaryDirectory() as tmp:
            workbook = Path(tmp) / "report.xlsx"
            comparison_json = Path(tmp) / "comparison_data.json"
            output = Path(tmp) / "sber_vs.html"
            history = {
                "scans": [{
                    "date": "2026-07-13T00:00:00",
                    "results": {
                        "alfa_only_1": {
                            "bank": "Альфа-Банк",
                            "tier": "Alfa Only — уровень 1",
                            "segment": "0–3 млн ₽",
                            "scan_date": "2026-07-13T00:00:00",
                            "sources_ok": 1,
                            "source_url": "https://example.test",
                            "status": "ok",
                            "score": {"total": 1, "breakdown": {}},
                            "fields": {
                                "entry_conditions": merged("2990 ₽ в мес"),
                                "service_cost": merged("2990 ₽ в мес"),
                                "lounge_access": merged("2 в мес (12 в год), доступ через | Mir Pass | ( | ·ON·PASS | ), | Phoenix Pass | , | Soft travel | , Only Assist"),
                                "cashback": merged("7% в 5 категориях | лимит 30 000 ₽"),
                                "deposits": merged(NOT_FOUND),
                                "taxi": merged(NOT_FOUND),
                                "restaurants": merged(NOT_FOUND),
                                "insurance": merged(NOT_FOUND),
                                "concierge": merged(NOT_FOUND),
                                "other_benefits": merged(NOT_FOUND),
                            },
                        },
                    },
                    "meta": {},
                }],
                "changelog": [],
            }
            write_report(history, workbook)
            write_comparison_json(history, comparison_json)
            build_sber_vs_landing(comparison_json, output)

            wb = load_workbook(workbook, read_only=True, data_only=True)
            visible_sheets = ["Сводная", "Альфа-Банк"]
            for sheet in visible_sheets:
                for row in wb[sheet].iter_rows(values_only=True):
                    for value in row:
                        self.assertNotIn("|", str(value or ""))
            html = output.read_text(encoding="utf-8")
            payload = re.search(
                r'<script id="data" type="application/json">(.*?)</script>',
                html,
                flags=re.S,
            ).group(1)
            self.assertNotIn("|", payload)

    def test_no_pipe_in_any_user_visible_field(self):
        with tempfile.TemporaryDirectory() as tmp:
            workbook, output = self._build_current_outputs(tmp)
            offenders = [
                (ctx, text[:200])
                for ctx, text in self._user_visible_texts(workbook, output)
                if "|" in text
            ]
        self.assertEqual(offenders, [])

    def test_no_truncated_word_in_any_bank(self):
        pattern = re.compile(r"[A-Za-zА-Яа-яЁё]+(?:\.\.\.|…)")
        broken_promo = re.compile(r"промо-ме(?:\s|$|[.;,])")
        with tempfile.TemporaryDirectory() as tmp:
            workbook, output = self._build_current_outputs(tmp)
            offenders = [
                (ctx, text[:200])
                for ctx, text in self._user_visible_texts(workbook, output)
                if pattern.search(text) or broken_promo.search(text) or "исследова..." in text
            ]
        self.assertEqual(offenders, [])

    def test_no_invalid_terminal_character(self):
        bad = ("-", "/", "|", "(", ",", ":")
        with tempfile.TemporaryDirectory() as tmp:
            workbook, output = self._build_current_outputs(tmp)
            offenders = [
                (ctx, text[-80:])
                for ctx, text in self._user_visible_texts(workbook, output)
                if not text.strip().startswith(("http://", "https://"))
                and text.strip().endswith(bad)
            ]
        self.assertEqual(offenders, [])

    def test_balanced_parentheses_all_banks(self):
        with tempfile.TemporaryDirectory() as tmp:
            workbook, output = self._build_current_outputs(tmp)
            offenders = [
                (ctx, text[:200])
                for ctx, text in self._user_visible_texts(workbook, output)
                if text.count("(") != text.count(")")
            ]
        self.assertEqual(offenders, [])

    def test_no_null_tokens_in_user_output(self):
        pattern = re.compile(r"\b(?:null|none|nan)\b", re.IGNORECASE)
        with tempfile.TemporaryDirectory() as tmp:
            workbook, output = self._build_current_outputs(tmp)
            offenders = [
                (ctx, text[:200])
                for ctx, text in self._user_visible_texts(workbook, output)
                if pattern.search(text)
            ]
        self.assertEqual(offenders, [])

    def test_no_double_spaces_all_banks(self):
        with tempfile.TemporaryDirectory() as tmp:
            workbook, output = self._build_current_outputs(tmp)
            offenders = [
                (ctx, text[:200])
                for ctx, text in self._user_visible_texts(workbook, output)
                if "  " in text
            ]
        self.assertEqual(offenders, [])

    def test_natural_list_formatting_all_banks(self):
        formatted = normalize_user_text(
            "| Mir Pass | ( | ON·PASS | ), | Phoenix Pass |, | Soft Travel |, Only Assist"
        )
        self.assertEqual(
            formatted,
            "Mir Pass, ON·PASS, Phoenix Pass, Soft Travel и Only Assist",
        )

    def test_full_details_preserved_all_banks(self):
        with tempfile.TemporaryDirectory() as tmp:
            _, output = self._build_current_outputs(tmp)
            html = output.read_text(encoding="utf-8")
            payload = re.search(
                r'<script id="data" type="application/json">(.*?)</script>',
                html,
                flags=re.S,
            ).group(1)
            data = self._comparison_payload(payload)
            details = [
                attr.get("details", "")
                for bank in data
                for level in bank["levels"]
                for attr in level["attrs"]
                if attr.get("details")
            ]
        self.assertGreater(len(details), 0)
        self.assertTrue(any(len(text) > 120 for text in details))

    def test_html_and_excel_use_same_formatter(self):
        raw = "2 в мес (12 в год), доступ через | Mir Pass | ( | ON·PASS | ), | Phoenix Pass |"
        self.assertEqual(normalize_user_text(raw), normalize_source_text(raw))

    def test_sber_new_label_and_score_hidden_in_html(self):
        with tempfile.TemporaryDirectory() as tmp:
            workbook = Path(tmp) / "report.xlsx"
            comparison_json = Path(tmp) / "comparison_data.json"
            output = Path(tmp) / "sber_vs.html"
            history = {
                "scans": [{
                    "date": "2026-07-15T00:00:00",
                    "results": {
                        "sber_premier_1": {
                            "bank": "Сбер",
                            "tier": "СберПремьер — уровень 1",
                            "segment": "0–3 млн ₽",
                            "scan_date": "2026-07-15T00:00:00",
                            "sources_ok": 1,
                            "source_url": "https://premiumbanking.info/sber/1",
                            "status": "ok",
                            "score": {"total": 4.8, "breakdown": {}},
                            "fields": self._comparison_fields(
                                entry_conditions="Новый СберПремьер за 2 млн ₽",
                                deposits="до 15% годовых",
                                restaurants="2 в мес по 5000 ₽",
                            ),
                        },
                        "alfa_only_1": {
                            "bank": "Альфа-Банк",
                            "tier": "Alfa Only — уровень 1",
                            "segment": "0–3 млн ₽",
                            "scan_date": "2026-07-15T00:00:00",
                            "sources_ok": 1,
                            "source_url": "https://premiumbanking.info/alfabank/1",
                            "status": "ok",
                            "score": {"total": 1.0, "breakdown": {}},
                            "fields": self._comparison_fields(
                                entry_conditions="Alfa Only за 3 млн ₽",
                                deposits="до 13% годовых",
                                restaurants="2 в мес по 2500 ₽",
                            ),
                        },
                    },
                    "meta": {},
                }],
                "changelog": [],
            }
            write_report(history, workbook)
            write_comparison_json(history, comparison_json)
            with patch("landing.sber_vs.premium_changes.load_changes", return_value=[]):
                build_sber_vs_landing(comparison_json, output)
            html = output.read_text(encoding="utf-8")
            payload = re.search(
                r'<script id="data" type="application/json">(.*?)</script>',
                html,
                flags=re.S,
            ).group(1)
            data = self._comparison_payload(payload)

        self.assertNotIn("Новый Сбер", html)
        self.assertNotIn("итоговый балл 0–5", html)
        self.assertNotIn("4.80", html)
        sber = next(bank for bank in data if bank["bank"] == "Сбер")["levels"][0]
        alfa = next(bank for bank in data if bank["bank"] == "Альфа-Банк")["levels"][0]
        sber_deposits = next(attr for attr in sber["attrs"] if attr["id"] == "deposits")
        alfa_deposits = next(attr for attr in alfa["attrs"] if attr["id"] == "deposits")
        sber_restaurants = next(attr for attr in sber["attrs"] if attr["id"] == "restaurants")
        alfa_restaurants = next(attr for attr in alfa["attrs"] if attr["id"] == "restaurants")

        self.assertGreater(sber_deposits["score"], alfa_deposits["score"])
        self.assertGreater(sber_restaurants["score"], alfa_restaurants["score"])

    def _comparison_fields(self, **overrides):
        fields = {
            "entry_conditions": NOT_FOUND,
            "service_cost": NOT_FOUND,
            "lounge_access": NOT_FOUND,
            "cashback": NOT_FOUND,
            "transfers_payments": NOT_FOUND,
            "cash_withdrawal": NOT_FOUND,
            "supreme": NOT_FOUND,
            "deposits": NOT_FOUND,
            "taxi": NOT_FOUND,
            "restaurants": NOT_FOUND,
            "insurance": NOT_FOUND,
            "concierge": NOT_FOUND,
            "other_benefits": NOT_FOUND,
        }
        fields.update(overrides)
        return {key: merged(value) for key, value in fields.items()}

    def test_all_current_banks_pass_text_quality_validation(self):
        required_banks = {
            "Сбер", "Альфа-Банк", "ВТБ", "Газпромбанк",
            "Озон Банк", "Райффайзен Банк", "Т-Банк",
        }
        with tempfile.TemporaryDirectory() as tmp:
            workbook, output = self._build_current_outputs(tmp)
            seen_banks = set()
            problems = []
            for ctx, text in self._user_visible_texts(workbook, output):
                for bank in required_banks:
                    if bank in text or bank in ctx:
                        seen_banks.add(bank)
                issues = validate_user_visible_text(text)
                if issues:
                    problems.append((ctx, issues, text[:200]))
        self.assertEqual(required_banks - seen_banks, set())
        self.assertEqual(problems, [])

    def test_no_corrupted_symbols_in_any_excel_sheet(self):
        with tempfile.TemporaryDirectory() as tmp:
            workbook, _ = self._build_current_outputs(tmp)
            wb = load_workbook(workbook, read_only=True, data_only=True)
            problems = []
            for sheet in wb.sheetnames:
                for row in wb[sheet].iter_rows():
                    for cell in row:
                        value = cell.value
                        if not isinstance(value, str) or not value.strip():
                            continue
                        issues = [
                            issue for issue in validate_user_visible_text(value)
                            if issue in {
                                "contains replacement character",
                                "looks like binary or corrupted text",
                            }
                        ]
                        if issues:
                            problems.append((sheet, cell.coordinate, issues, value[:120]))
        self.assertEqual(problems, [])


if __name__ == "__main__":
    unittest.main()
