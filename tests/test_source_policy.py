import inspect
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from landing import premium_changes
from landing import sber_vs
from landing.premium_changes import build_premium_changes_landing
from landing.sber_vs import build_sber_vs_landing
from report.json_writer import write_comparison_json
from scanner.benefits import build_other_benefits
from scanner.contracts import validate_scan_contracts
from scanner.diff import change_kind, diff_results
from scanner.merge import merge_tier_fields
from scanner.publication import apply_publication_gate
from scanner.publication import derivation_components, gate_field
from scanner.sources import (
    AUTHORITATIVE_SOURCE_URLS,
    NOT_FOUND,
    NOT_FOUND_AVAILABLE,
    REQUIRED_PRIORITY_URLS,
    SOURCE_META,
    SOURCE_PRIORITY_ORDER,
    get_bank,
    registered_source_urls,
    is_authoritative_url,
    tier_sources,
)


def parsed(source_id, value, url=None, quality="snippet", field_id="cashback"):
    return {
        "source_id": source_id,
        "url": url or f"https://example.com/{source_id}",
        "quality": quality,
        "fields": {field_id: value},
    }


class SourcePolicyTests(unittest.TestCase):
    def test_official_source_has_priority(self):
        merged = merge_tier_fields(
            [
                parsed("pbi", "PBI: 10 визитов", "https://premiumbanking.info/sber/1", "structured"),
                parsed("official", "Official: 4 визита", "https://bank.example/tariff.pdf", "structured"),
            ],
            {},
            ["cashback"],
            "2026-07-13T10:00:00",
            bank_id="sber",
            tier_id="sber_premier_1",
        )

        field = merged["cashback"]
        self.assertEqual(field["value"], "Official: 4 визита")
        self.assertEqual(field["source_type"], "official")
        self.assertTrue(field["divergent"])

    def test_structured_pbi_beats_official_keyword_snippet(self):
        merged = merge_tier_fields(
            [
                parsed("pbi", "PBI: 10 визитов", "https://premiumbanking.info/sber/1", "structured"),
                parsed("official", "Official snippet: 4 визита", "https://bank.example/tariff.pdf", "snippet"),
            ],
            {},
            ["cashback"],
            "2026-07-13T10:00:00",
            bank_id="sber",
            tier_id="sber_premier_1",
        )

        field = merged["cashback"]
        self.assertEqual(field["value"], "PBI: 10 визитов")
        self.assertEqual(field["source_type"], "premiumbanking.info")

    def test_pbi_used_when_official_field_missing(self):
        merged = merge_tier_fields(
            [
                parsed("official", NOT_FOUND, "https://bank.example/tariff.pdf"),
                parsed("pbi", "PBI value", "https://premiumbanking.info/sber/1", "structured"),
            ],
            {},
            ["cashback"],
            "2026-07-13T10:00:00",
            bank_id="sber",
            tier_id="sber_premier_1",
        )

        self.assertEqual(merged["cashback"]["value"], "PBI value")
        self.assertEqual(merged["cashback"]["source_type"], "premiumbanking.info")

    def test_pbi_used_when_official_value_has_broken_encoding(self):
        merged = merge_tier_fields(
            [
                parsed("official", "Ð¡ÐµÑÐ²Ð¸Ñ Mir Pass", "https://bank.example/tariff.pdf",
                       field_id="lounge_access"),
                parsed("pbi", "2 в мес, 24 в год", "https://premiumbanking.info/vtb",
                       "structured", field_id="lounge_access"),
            ],
            {},
            ["lounge_access"],
            "2026-07-15T10:00:00",
            bank_id="vtb",
            tier_id="vtb_privilege_2",
        )

        self.assertEqual(merged["lounge_access"]["value"], "2 в мес, 24 в год")
        self.assertEqual(merged["lounge_access"]["source_type"], "premiumbanking.info")
        self.assertNotIn("Ð", merged["lounge_access"]["value"])

    def test_pbi_used_when_official_source_unavailable(self):
        merged = merge_tier_fields(
            [parsed("pbi", "PBI fallback", "https://premiumbanking.info/tbank/1", "structured")],
            {},
            ["cashback"],
            "2026-07-13T10:00:00",
            bank_id="tbank",
            tier_id="tbank_bronze",
        )

        self.assertEqual(merged["cashback"]["value"], "PBI fallback")
        self.assertEqual(merged["cashback"]["source_type"], "premiumbanking.info")

    def test_pbi_level_beats_general_pbi_page_for_same_quality(self):
        merged = merge_tier_fields(
            [
                parsed("pbi", "Короткая общая страница", "https://premiumbanking.info/gazprombank",
                       "structured", field_id="ecosystem"),
                parsed("pbi_level", "Полная страница уровня", "https://premiumbanking.info/gazprombank/4",
                       "structured", field_id="ecosystem"),
            ],
            {},
            ["ecosystem"],
            "2026-07-15T10:00:00",
            bank_id="gazprombank",
            tier_id="gpb_private",
        )

        self.assertEqual(merged["ecosystem"]["value"], "Полная страница уровня")
        self.assertEqual(merged["ecosystem"]["source_id"], "pbi_level")

    def test_publication_gate_blocks_generic_snippet_for_html(self):
        merged = merge_tier_fields(
            [parsed("official", "Official snippet value", "https://bank.example/tariff.pdf")],
            {},
            ["cashback"],
            "2026-07-15T10:00:00",
            bank_id="sber",
            tier_id="sber_premier_1",
        )

        gated = apply_publication_gate(merged)
        field = gated["cashback"]

        self.assertEqual(field["value"], NOT_FOUND)
        self.assertEqual(field["publication_status"], "blocked")
        self.assertEqual(field["blocked_value"], "Official snippet value")
        self.assertIn("snippet", field["publication_reason"])

    def test_publication_gate_allows_structured_pbi(self):
        merged = merge_tier_fields(
            [parsed("pbi", "2 визита в месяц", "https://premiumbanking.info/sber/1",
                    "structured", "lounge_access")],
            {},
            ["lounge_access"],
            "2026-07-15T10:00:00",
            bank_id="sber",
            tier_id="sber_premier_1",
        )

        gated = apply_publication_gate(merged)

        self.assertEqual(gated["lounge_access"]["value"], "2 визита в месяц")
        self.assertEqual(gated["lounge_access"]["publication_status"], "published")

    def test_publication_gate_does_not_block_structured_pbi_by_generic_snippet(self):
        merged = merge_tier_fields(
            [
                parsed(
                    "pbi",
                    "опция «Такси и рестораны» — 1 такси на 1000 ₽, 1 ресторан на 2000 ₽",
                    "https://premiumbanking.info/sber",
                    "structured",
                    "taxi",
                ),
                parsed(
                    "sravni_ru",
                    "Кафе, рестораны, такси | За посещение ресторана (~ 6 раз в месяц)",
                    "https://www.sravni.ru/enciklopediya/info/sberbank-premer-chto-ehto-takoe/",
                    "snippet",
                    "taxi",
                ),
            ],
            {},
            ["taxi"],
            "2026-07-15T10:00:00",
            bank_id="sber",
            tier_id="sber_premier_2",
        )

        self.assertTrue(merged["taxi"]["divergent"])
        gated = apply_publication_gate(merged)

        self.assertEqual(
            gated["taxi"]["value"],
            "опция «Такси и рестораны» — 1 такси на 1000 ₽, 1 ресторан на 2000 ₽",
        )
        self.assertEqual(gated["taxi"]["publication_status"], "published")

    def test_authoritative_source_not_blocked_by_non_authoritative_conflict(self):
        merged = merge_tier_fields(
            [
                parsed(
                    "pbi",
                    "4 визита",
                    "https://premiumbanking.info/vtb",
                    "structured",
                    "lounge_access",
                ),
                parsed(
                    "frankrg",
                    "8 визитов",
                    "https://example.com/non-authoritative",
                    "structured",
                    "lounge_access",
                ),
            ],
            {},
            ["lounge_access"],
            "2026-07-15T10:00:00",
            bank_id="vtb",
            tier_id="vtb_privilege_1",
        )

        self.assertTrue(merged["lounge_access"]["divergent"])
        gated = apply_publication_gate(merged)

        self.assertEqual(gated["lounge_access"]["value"], "4 визита")
        self.assertEqual(gated["lounge_access"]["publication_status"], "published")

    def test_derived_other_benefits_ignores_not_found_blocked_components(self):
        fields = {
            "always_included_options": {
                "value": NOT_FOUND,
                "source_id": "sravni_ru",
                "source_url": "https://example.com/noisy",
                "quality": "snippet",
                "publication_status": "blocked",
            },
            "selectable_options": {
                "value": "опция «Самокат» — 2 заказа по 500 ₽",
                "source_id": "pbi",
                "source_url": "https://premiumbanking.info/sber",
                "quality": "structured",
                "publication_status": "published",
            },
        }
        derived = {
            "value": "• Самокат — 2 заказа по 500 ₽ [опция на выбор]",
            "source_id": "derived",
            "source_url": "https://premiumbanking.info/sber",
            "quality": "derived",
            "raw_text": "• Самокат — 2 заказа по 500 ₽ [опция на выбор]",
            "derived_from": derivation_components(
                fields, ("always_included_options", "selectable_options")
            ),
        }

        gated = gate_field(derived, "other_benefits")

        self.assertEqual(gated["publication_status"], "published")
        self.assertEqual(
            [item["field_id"] for item in gated["derived_from"]],
            ["selectable_options"],
        )

    def test_excel_and_html_hide_blocked_snippet_but_keep_provenance(self):
        with tempfile.TemporaryDirectory() as tmp:
            workbook = Path(tmp) / "report.xlsx"
            comparison_json = Path(tmp) / "comparison_data.json"
            output = Path(tmp) / "sber_vs.html"
            merged = merge_tier_fields(
                [parsed("official", "False snippet", "https://bank.example/tariff.pdf")],
                {},
                ["cashback"],
                "2026-07-15T10:00:00",
                bank_id="sber",
                tier_id="sber_premier_1",
            )
            fields = apply_publication_gate(merged)
            history = {
                "scans": [{
                    "date": "2026-07-15T10:00:00",
                    "results": {
                        "sber_premier_1": {
                            "bank": "Сбер",
                            "tier": "СберПремьер — уровень 1",
                            "segment": "0–3 млн ₽",
                            "fields": fields,
                            "source_url": "https://bank.example/tariff.pdf",
                            "status": "ok",
                            "sources_ok": 1,
                            "scan_date": "2026-07-15T10:00:00",
                            "score": {"total": 0, "breakdown": {}},
                        },
                    },
                    "meta": {"quality_issues": []},
                }],
                "changelog": [],
            }

            from report.excel_writer import write_report
            write_report(history, workbook)
            write_comparison_json(history, comparison_json)
            with patch.object(sber_vs.premium_changes, "load_changes", return_value=[]):
                build_sber_vs_landing(comparison_json, output)

            wb = load_workbook(workbook, read_only=True, data_only=True)
            summary_rows = list(wb["Сводная"].iter_rows(values_only=True))
            provenance_rows = list(wb["Провенанс значений"].iter_rows(values_only=True))
            passport_rows = list(wb["Source_Passport"].iter_rows(values_only=True))
            html = output.read_text(encoding="utf-8")

        self.assertTrue(any(NOT_FOUND_AVAILABLE in row for row in summary_rows if row))
        self.assertIn(NOT_FOUND_AVAILABLE, html)
        self.assertNotIn("False snippet", html)
        self.assertTrue(any("False snippet" in row for row in provenance_rows if row))
        self.assertTrue(any(
            isinstance(cell, str) and "Кэшбэк" in cell
            for row in passport_rows
            for cell in row
        ))

    def test_quality_contract_flags_corrupted_text_and_missing_vtb_markers(self):
        results = {
            "vtb_privilege_2": {
                "bank": "ВТБ",
                "tier": "Привилегия — уровень 2",
                "fields": {
                    "entry_conditions": {"value": "за 2.5 млн ₽ для Мск", "source_url": "https://premiumbanking.info/vtb", "raw_text": "за 2.5 млн ₽"},
                    "selection_rules": {"value": "Преференции: 2 в мес", "source_url": "https://premiumbanking.info/vtb", "raw_text": "Преференции: 2 в мес"},
                    "lounge_access": {"value": "Ð¡ÐµÑÐ²Ð¸Ñ Mir Pass", "source_url": "https://vtb.example", "raw_text": "Ð¡ÐµÑÐ²Ð¸Ñ Mir Pass"},
                    "restaurants": {"value": "2 в мес (12 в год) по 2500 ₽", "source_url": "https://premiumbanking.info/vtb", "raw_text": "2 в мес"},
                    "taxi": {"value": "2 в мес, 12 в год (на 1000 ₽)", "source_url": "https://premiumbanking.info/vtb", "raw_text": "2 в мес"},
                    "insurance": {"value": "$100/100 тыс", "source_url": "https://premiumbanking.info/vtb", "raw_text": "$100/100 тыс"},
                    "concierge": {"value": "АМА консьерж", "source_url": "https://premiumbanking.info/vtb", "raw_text": "АМА консьерж"},
                    "other_benefits": {"value": "Помощь на дорогах", "source_id": "derived", "raw_text": "Помощь на дорогах"},
                },
            }
        }

        issues = validate_scan_contracts(results)
        codes = {issue["code"] for issue in issues}
        self.assertIn("text_quality", codes)
        self.assertIn("field_marker", codes)
        self.assertTrue(any(issue["severity"] == "error" for issue in issues))

    def test_no_value_invented_when_sources_empty(self):
        merged = merge_tier_fields(
            [],
            {},
            ["cashback"],
            "2026-07-13T10:00:00",
            bank_id="sber",
            tier_id="sber_premier_1",
        )

        field = merged["cashback"]
        self.assertEqual(field["value"], NOT_FOUND)
        self.assertEqual(field["conflict_status"], "not_found")
        self.assertEqual(field["raw_text"], "")

    def test_every_fact_has_source_url(self):
        merged = merge_tier_fields(
            [parsed("official", "Official value", "https://bank.example/tariff.pdf")],
            {},
            ["cashback"],
            "2026-07-13T10:00:00",
            bank_id="sber",
            tier_id="sber_premier_1",
        )

        self.assertEqual(merged["cashback"]["source_url"], "https://bank.example/tariff.pdf")

    def test_every_fact_has_tier_id(self):
        merged = merge_tier_fields(
            [parsed("official", "Official value")],
            {},
            ["cashback"],
            "2026-07-13T10:00:00",
            bank_id="sber",
            tier_id="sber_premier_1",
        )

        self.assertEqual(merged["cashback"]["bank_id"], "sber")
        self.assertEqual(merged["cashback"]["tier_id"], "sber_premier_1")
        self.assertEqual(merged["cashback"]["field_id"], "cashback")

    def test_no_cross_tier_fallback(self):
        tier_1 = merge_tier_fields(
            [parsed("official", "Tier 1 value")],
            {},
            ["cashback"],
            "2026-07-13T10:00:00",
            bank_id="sber",
            tier_id="sber_premier_1",
        )
        tier_2 = merge_tier_fields(
            [],
            {},
            ["cashback"],
            "2026-07-13T10:00:00",
            bank_id="sber",
            tier_id="sber_premier_2",
        )

        self.assertEqual(tier_1["cashback"]["value"], "Tier 1 value")
        self.assertEqual(tier_2["cashback"]["value"], NOT_FOUND)

    def test_no_cross_bank_fallback(self):
        sber = merge_tier_fields(
            [parsed("official", "Sber value")],
            {},
            ["cashback"],
            "2026-07-13T10:00:00",
            bank_id="sber",
            tier_id="sber_premier_1",
        )
        alfa = merge_tier_fields(
            [],
            {},
            ["cashback"],
            "2026-07-13T10:00:00",
            bank_id="alfa",
            tier_id="alfa_only_1",
        )

        self.assertEqual(sber["cashback"]["value"], "Sber value")
        self.assertEqual(alfa["cashback"]["value"], NOT_FOUND)

    def test_html_uses_json_only(self):
        source = inspect.getsource(sber_vs)
        self.assertNotIn("requests.get", source)
        self.assertNotIn("Fetcher(", source)
        self.assertNotIn("load_workbook", source)

        with tempfile.TemporaryDirectory() as tmp:
            comparison_json = Path(tmp) / "comparison_data.json"
            output = Path(tmp) / "sber_vs.html"
            history = {
                "scans": [{
                    "date": "2026-07-13T10:00:00",
                    "results": {
                        "sber_premier_1": {
                            "bank": "Сбер",
                            "tier": "СберПремьер — уровень 1",
                            "segment": "0–3 млн ₽",
                            "scan_date": "2026-07-13T10:00:00",
                            "sources_ok": 1,
                            "source_url": "https://bank.example/source.pdf",
                            "status": "ok",
                            "score": {"total": 1.0, "breakdown": {}},
                            "fields": {
                                "entry_conditions": {
                                    "value": "JSON-only condition",
                                    "source_url": "https://bank.example/source.pdf",
                                    "raw_text": "JSON-only condition",
                                    "source_id": "official",
                                    "source_type": "official",
                                    "publication_status": "published",
                                },
                            },
                        },
                    },
                    "meta": {},
                }],
                "changelog": [],
            }
            write_comparison_json(history, comparison_json)

            with patch.object(sber_vs.premium_changes, "load_changes", return_value=[]):
                build_sber_vs_landing(comparison_json, output)
            html = output.read_text(encoding="utf-8")

        self.assertIn("JSON-only condition", html)

    def test_excel_report_contains_quality_issues_sheet(self):
        with tempfile.TemporaryDirectory() as tmp:
            workbook = Path(tmp) / "report.xlsx"
            history = {
                "scans": [{
                    "date": "2026-07-15T10:00:00",
                    "results": {},
                    "meta": {
                        "quality_issues": [{
                            "severity": "error",
                            "code": "field_marker",
                            "bank": "ВТБ",
                            "tier": "Привилегия — уровень 2",
                            "tier_id": "vtb_privilege_2",
                            "field_id": "lounge_access",
                            "message": "Не найден обязательный маркер: 24 в год",
                            "value": "2 в мес",
                        }]
                    },
                }],
                "changelog": [],
            }

            from report.excel_writer import write_report
            write_report(history, workbook)
            wb = load_workbook(workbook, read_only=True, data_only=True)
            sheetnames = wb.sheetnames
            rows = list(wb["Quality_Issues"].iter_rows(values_only=True))

        self.assertIn("Quality_Issues", sheetnames)
        self.assertIn(("error", "field_marker", "ВТБ", "Привилегия — уровень 2",
                       "vtb_privilege_2", "lounge_access",
                       "Не найден обязательный маркер: до 24 в год",
                       "2 посещения в месяц"), rows)

    def test_premium_changes_html_uses_pbi_updates_only(self):
        source = inspect.getsource(premium_changes)
        self.assertIn("requests.get", source)
        self.assertNotIn("Fetcher(", source)
        self.assertIn("https://premiumbanking.info/sber", source)
        self.assertNotIn("load_workbook", source)

        with tempfile.TemporaryDirectory() as tmp:
            workbook = Path(tmp) / "report.xlsx"
            output = Path(tmp) / "premium_changes.html"
            workbook.write_text("not used", encoding="utf-8")
            records = [{
                "bank": "Сбер",
                "dateLabel": "апрель 2026",
                "dateSort": "2026-04-01",
                "text": "В Сбере опубликована новость из ПБИ.",
                "sourcePage": "https://premiumbanking.info/sber",
                "order": 1,
            }]

            with patch.object(premium_changes, "fetch_pbi_updates", return_value=(records, 0)):
                build_premium_changes_landing(workbook, output)
            html = output.read_text(encoding="utf-8")

        self.assertIn("Последние изменения", html)
        self.assertIn("В Сбере опубликована новость из ПБИ.", html)
        self.assertIn("Сбер", html)
        self.assertIn(
            "https://www.sberbank.ru/common/img/uploaded/files/pdf/tarif_premobsl_06032026.pdf",
            html,
        )
        self.assertNotIn('href="https://premiumbanking.info/sber"', html)
        self.assertNotIn("Было", html)
        self.assertNotIn("Стало", html)
        self.assertNotIn("Подтверждено", html)

    def test_changes_panel_has_only_bottom_hide_button(self):
        html = premium_changes.render_changes_panel(
            [{"name": "Сбер", "changes": []}],
            __import__("datetime").datetime(2026, 7, 15, 12, 0),
        )

        self.assertEqual(html.count("Скрыть изменения"), 1)
        self.assertNotIn("changes-panel-head", html)
        self.assertIn("changes-sticky-close", html)

    def test_premium_changes_parser_reads_only_updates_block(self):
        page = """
        <span id="updates"></span>
        <div class="container">
          <p class="h2"><strong>Последние изменения премиальной программы в Сбере</strong></p>
          <p>Первая новость. <a href="/journal/x">Подробнее</a>.</p>
          <footer>апр&nbsp;2026</footer>
          <p>Вторая новость.</p>
          <footer>март&nbsp;2026</footer>
        </div>
        <div><span id="all_levels"></span><p>Таблица уровней не нужна</p></div>
        """

        records = premium_changes.parse_pbi_updates(
            page, "Сбер", "https://premiumbanking.info/sber")

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0]["bank"], "Сбер")
        self.assertEqual(records[0]["dateLabel"], "апр 2026")
        self.assertEqual(records[0]["dateSort"], "2026-04-01")
        self.assertEqual(records[0]["text"], "Первая новость.")
        self.assertEqual(records[0]["sourcePage"], "https://premiumbanking.info/sber")
        self.assertEqual(records[0]["order"], 1)
        self.assertNotIn("Таблица уровней", " ".join(item["text"] for item in records))

    def test_reference_fields_do_not_create_changelog(self):
        prev = {
            "date": "2026-07-15T10:00:00",
            "results": {
                "alfa_aclub": {
                    "bank": "Альфа-Банк",
                    "tier": "A-Club (private)",
                    "fields": {"other_notes": {"value": NOT_FOUND}},
                },
            },
        }
        new = {
            "date": "2026-07-15T11:00:00",
            "results": {
                "alfa_aclub": {
                    "bank": "Альфа-Банк",
                    "tier": "A-Club (private)",
                    "fields": {
                        "other_notes": {
                            "value": "Рейтинг уровня по отзывам ПБИ: 7 / 10"
                        },
                    },
                },
            },
        }

        self.assertEqual(diff_results(prev, new, {"other_notes": "Прочее"}), [])

    def test_workbook_contains_monitoring_contract_sheets(self):
        with tempfile.TemporaryDirectory() as tmp:
            workbook = Path(tmp) / "report.xlsx"
            history = {
                "scans": [{
                    "date": "2026-07-14T10:00:00",
                    "results": {},
                    "meta": {"mode": "all", "sources_ok": [], "sources_failed": {}},
                }],
                "changelog": [],
            }
            from report.excel_writer import write_report

            write_report(history, workbook)
            wb = load_workbook(workbook, read_only=True, data_only=True)

        for sheet in ("Banks", "Products", "Changes", "Sources", "Monitoring_Log"):
            self.assertIn(sheet, wb.sheetnames)

    def test_changes_contract_preserves_absolute_source_url(self):
        with tempfile.TemporaryDirectory() as tmp:
            workbook = Path(tmp) / "report.xlsx"
            history = {
                "scans": [{
                    "date": "2026-07-14T10:00:00",
                    "results": {},
                    "meta": {"mode": "all", "sources_ok": [], "sources_failed": {}},
                }],
                "changelog": [{
                    "scan_date": "2026-07-14T10:00:00",
                    "prev_date": "2026-07-13T10:00:00",
                    "bank": "Сбер",
                    "tier": "СберПремьер — уровень 1",
                    "field": "Бизнес-залы (визиты, спутники)",
                    "old": "2 посещения",
                    "new": "4 посещения",
                    "kind": "market",
                    "source": "premiumbanking.info",
                    "source_url": "https://premiumbanking.info/sber/1",
                }],
            }
            from report.excel_writer import write_report

            write_report(history, workbook)
            wb = load_workbook(workbook, read_only=True, data_only=True)
            ws = wb["Changes"]
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            source_url_col = headers.index("source_url") + 1

            self.assertEqual(
                ws.cell(row=2, column=source_url_col).value,
                "https://premiumbanking.info/sber/1",
            )

    def test_source_conflict_is_logged(self):
        merged = merge_tier_fields(
            [
                parsed("official", "Official 4 визита", "https://bank.example/tariff.pdf", "structured"),
                parsed("pbi", "PBI 10 визитов", "https://premiumbanking.info/sber/1", "structured"),
            ],
            {},
            ["cashback"],
            "2026-07-13T10:00:00",
            bank_id="sber",
            tier_id="sber_premier_1",
        )

        field = merged["cashback"]
        self.assertEqual(field["conflict_status"], "conflict")
        self.assertTrue(field["alternatives"])
        self.assertEqual(field["alternatives"][0]["source_type"], "premiumbanking.info")

    def test_unknown_status_for_ambiguous_fact(self):
        merged = merge_tier_fields(
            [parsed("official", "может быть доступно по решению банка", quality="ambiguous")],
            {},
            ["cashback"],
            "2026-07-13T10:00:00",
            bank_id="sber",
            tier_id="sber_premier_1",
        )

        self.assertEqual(merged["cashback"]["conflict_status"], "unknown")
        self.assertEqual(merged["cashback"]["raw_text"], "может быть доступно по решению банка")

    def test_insurance_not_in_other_benefits(self):
        fields = {
            "always_included_options": {
                "value": "Здоровье — телемедицина, анализы, исследования",
                "source_url": "https://premiumbanking.info/sber/4",
                "date_checked": "2026-07-13",
                "source_name": "premiumbanking.info",
                "source_id": "pbi",
                "quality": "structured",
                "raw_text": "Здоровье — телемедицина, анализы, исследования",
            },
            "selectable_options": {"value": NOT_FOUND},
            "selection_rules": {"value": NOT_FOUND},
            "auto": {"value": NOT_FOUND},
            "ecosystem": {"value": NOT_FOUND},
            "insurance": {
                "value": "$1 млн, 90 дн, ассистанс Mondial Assistance, потеря багажа",
            },
        }

        benefits = build_other_benefits(fields)
        text = " ".join(
            f"{item['title']} {item.get('description', '')}"
            for item in benefits
        ).lower()
        self.assertIn("здоровье", text)
        self.assertNotIn("ассистанс", text)
        self.assertNotIn("потеря багажа", text)

    def test_checkup_medical_benefits_are_not_filtered_as_receipts(self):
        benefits = build_other_benefits({
            "ecosystem": {
                "value": (
                    "Медицинская программа от «Доктор рядом» с телемедициной, "
                    "2 чекапа на выбор в год ; "
                    "Сервис «Лучи»: медицинские онлайн-консультации, "
                    "10 посещений по ДМС «Лучи» на тарифе «Бизнес» + "
                    "компенсация чекапа за рубежом и чекап «Здоровый образ жизни»"
                )
            }
        })
        text = " ".join(
            f"{item['title']} {item.get('description', '')}"
            for item in benefits
        )

        self.assertIn("Доктор рядом", text)
        self.assertIn("2 чекапа", text)
        self.assertIn("Лучи", text)
        self.assertIn("Здоровый образ жизни", text)

    def test_colon_benefit_keeps_quoted_service_name(self):
        benefits = build_other_benefits({
            "ecosystem": {
                "value": "Сервис «Лучи»: медицинские онлайн-консультации"
            }
        })

        self.assertEqual(benefits[0]["title"], "Сервис «Лучи»")
        self.assertEqual(benefits[0]["description"], "медицинские онлайн-консультации")

    def test_all_priority_urls_registered(self):
        self.assertEqual(REQUIRED_PRIORITY_URLS - registered_source_urls(), set())

    def test_all_authoritative_urls_registered(self):
        self.assertEqual(AUTHORITATIVE_SOURCE_URLS - registered_source_urls(), set())
        self.assertTrue(is_authoritative_url("https://premiumbanking.info/sber"))
        self.assertTrue(is_authoritative_url("https://www.tbank.ru/tinkoff-premium/"))
        self.assertTrue(is_authoritative_url("https://www.sberbank.com/ru/person/sb_premier_new?ysclid=x"))

    def test_aclub_official_source_registered(self):
        alfa = get_bank("alfa")
        aclub = next(tier for tier in alfa["tiers"] if tier["tier_id"] == "alfa_aclub")
        official = next(src for src in aclub["sources"] if src["source_id"] == "official")

        self.assertEqual(official["urls"][0], "https://alfabank.ru/a-club/")
        self.assertIn("https://alfabank.ru/a-club/", REQUIRED_PRIORITY_URLS)
        self.assertIn("https://alfabank.ru/a-club/", registered_source_urls())

    def test_aclub_not_merged_with_alfa_only(self):
        from scanner.curated import curated_for

        facts = curated_for("alfa_aclub")
        forbidden = ("Alfa Only заказывается", "Only Assist", "Tariffs_Alfa_Only_Card")

        for field_id, fact in facts.items():
            combined = " ".join([
                fact.get("value", ""),
                fact.get("source_url", ""),
                fact.get("note", ""),
            ])
            for marker in forbidden:
                with self.subTest(field_id=field_id, marker=marker):
                    self.assertNotIn(marker, combined)

    def test_no_sber_status_hardcode_in_html(self):
        source = inspect.getsource(sber_vs)

        self.assertNotIn("sber_first_4", source)
        self.assertNotIn("sber_first_5", source)
        self.assertNotIn("sber_private_6", source)
        self.assertNotIn("hide_selectable_badge", source)

    def test_ozon_deposits_are_curated_from_official_source(self):
        from scanner.curated import curated_for

        fact = curated_for("ozonbank_ultra_bronze")["deposits"]

        self.assertIn("13,5%", fact["value"])
        self.assertIn("15,1%", fact["value"])
        self.assertEqual(
            fact["source_url"],
            "https://finance.ozon.ru/promo/deposit/landing",
        )

    def test_vtb_prime_cashback_and_deposits_are_official_curated(self):
        from scanner.curated import curated_for

        for tier_id in ("vtb_prime_5", "vtb_prime_8"):
            facts = curated_for(tier_id)
            cashback = facts["cashback"]
            deposits = facts["deposits"]

            self.assertIn("до 30 000 ₽", cashback["value"])
            self.assertIn("3 категории из 9", cashback["value"])
            self.assertEqual(
                cashback["source_url"],
                "https://www.vtb.ru/privilegia/karty/debetovye/"
                "privilegiya-mir-supreme/",
            )
            self.assertIn("до 13,6%", deposits["value"])
            self.assertEqual(
                deposits["source_url"],
                "https://www.vtb.ru/privilegia/",
            )
            self.assertEqual(cashback["date_checked"], "2026-07-15")
            self.assertEqual(deposits["date_checked"], "2026-07-15")

    def test_source_priority_order(self):
        self.assertEqual(SOURCE_PRIORITY_ORDER, ["curated", "official", "pbi", "other"])
        self.assertLess(SOURCE_META["official"]["priority"], SOURCE_META["pbi"]["priority"])
        self.assertLess(SOURCE_META["pbi"]["priority"], SOURCE_META["banki_ru"]["priority"])

        sber = get_bank("sber")
        sources = tier_sources(sber, sber["tiers"][0])
        priorities = [SOURCE_META[src["source_id"]]["priority"] for src in sources]
        self.assertEqual(priorities, sorted(priorities))

    def test_diff_hides_wording_only_lounge_change(self):
        self.assertEqual(
            change_kind("lounge_access", "безлимит", "Бизнес-залы — безлимит"),
            "service",
        )

    def test_diff_hides_wording_only_taxi_change(self):
        self.assertEqual(
            change_kind(
                "taxi",
                "3 посещения в месяц, до 15 в год (по 5000 ₽)",
                "Такси — 3 раза в месяц, 15 раз в год, до 5 000 ₽",
            ),
            "service",
        )

    def test_diff_hides_removed_explanatory_parenthetical(self):
        self.assertEqual(
            change_kind(
                "entry_requirements",
                "60 млн ₽ на счетах для Москвы; 30 млн ₽ для регионов "
                "(порог после 2024 года; ранее ~6 млн ₽ среднемесячного остатка "
                "давали доступ к части сервисов)",
                "60 млн ₽ на счетах для Москвы; 30 млн ₽ для регионов",
            ),
            "service",
        )

    def test_diff_hides_wording_only_concierge_change(self):
        self.assertEqual(
            change_kind(
                "concierge",
                "Есть — консьерж-сервис PRIME (входит в основные привилегии А-Клуба, бессрочно)",
                "Есть — консьерж-сервис PRIME",
            ),
            "service",
        )

    def test_diff_marks_real_numeric_limit_change(self):
        self.assertEqual(
            change_kind(
                "lounge_access",
                "Бизнес-залы — безлимит",
                "Бизнес-залы — 2 посещения в месяц",
            ),
            "market",
        )

    def test_diff_marks_real_concierge_provider_change(self):
        self.assertEqual(
            change_kind(
                "concierge",
                "Есть — консьерж PRIME",
                "Есть — консьерж Aspire",
            ),
            "market",
        )


if __name__ == "__main__":
    unittest.main()
