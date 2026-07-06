# -*- coding: utf-8 -*-
"""Excel sheets for Voice of Customer feedback intelligence."""

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FEEDBACK_SHEETS = [
    "Customer Feedback",
    "Sentiment",
    "Advantages",
    "Disadvantages",
    "Wishes",
    "Problems",
    "Recommendations",
    "Trends",
    "Source Statistics",
]

HEADER_FILL = PatternFill("solid", fgColor="385723")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(*[Side(style="thin", color="CCCCCC")] * 4)


def write_feedback_report(history: dict, output_path: Path) -> dict:
    wb = load_workbook(output_path) if output_path.exists() else Workbook()
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
        wb.remove(wb["Sheet"])
    for sheet in FEEDBACK_SHEETS:
        if sheet in wb.sheetnames:
            wb.remove(wb[sheet])
    scan = history["scans"][-1] if history.get("scans") else _empty_scan()
    scan = _with_merged_source_stats(scan, history)
    _write_customer_feedback(wb, scan)
    _write_sentiment(wb, scan)
    _write_ranked(wb, scan, "Advantages", "advantages")
    _write_ranked(wb, scan, "Disadvantages", "disadvantages")
    _write_ranked(wb, scan, "Wishes", "wishes")
    _write_problems(wb, scan)
    _write_recommendations(wb, scan)
    _write_trends(wb, scan)
    _write_source_stats(wb, scan)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return {"output": str(output_path), "reviews": len(scan.get("reviews", []))}


def _with_merged_source_stats(scan: dict, history: dict) -> dict:
    merged = {}
    for item in history.get("scans", []):
        for stat in item.get("meta", {}).get("source_stats", []):
            source_id = stat.get("source_id")
            if source_id:
                merged[source_id] = stat
    if not merged:
        return scan
    result = dict(scan)
    meta = dict(scan.get("meta", {}))
    meta["source_stats"] = list(merged.values())
    result["meta"] = meta
    return result


def _empty_scan():
    return {"reviews": [], "analyses": {}, "suggestions": [], "trends": {}, "insights": {}, "meta": {}}


def _headers(ws, headers):
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP
        cell.border = THIN_BORDER
    ws.freeze_panes = "A2"


def _widths(ws, widths):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = WRAP
    cell.border = THIN_BORDER
    return cell


def _write_customer_feedback(wb, scan):
    ws = wb.create_sheet("Customer Feedback")
    headers = ["Review ID", "Record Type", "Источник", "URL", "Дата публикации",
               "Дата скана", "Автор", "Рейтинг", "Название", "Полный текст",
               "Достоинства", "Недостатки", "Комментарии", "Лайки", "Продукт",
               "Sentiment", "Эмоции", "Темы", "Источник данных", "Provenance"]
    _headers(ws, headers)
    _widths(ws, [18, 14, 20, 42, 14, 17, 18, 8, 32, 75, 45, 45, 35, 8, 18, 18, 24, 35, 20, 60])
    analyses = scan.get("analyses", {})
    for row, review in enumerate(scan.get("reviews", []), start=2):
        analysis = analyses.get(review["review_id"], {})
        prov = review.get("provenance", {})
        values = [
            review["review_id"], review.get("record_type", ""), review.get("source_name", ""),
            review.get("url", ""), review.get("published_at", review.get("date", "")),
            review.get("scanned_at", review.get("collected_at", "")), review.get("author", ""),
            review.get("rating"), review.get("title", ""), review.get("full_text", review.get("text", "")),
            review.get("pros", ""), review.get("cons", ""), review.get("comments", ""),
            review.get("likes_count"), review.get("product_id", ""),
            analysis.get("sentiment", ""), "; ".join(analysis.get("emotions", [])),
            "; ".join(analysis.get("topics", [])), review.get("data_source", ""),
            "; ".join(f"{k}: {v}" for k, v in prov.items() if v not in ("", None)),
        ]
        for col, value in enumerate(values, start=1):
            _cell(ws, row, col, value)


def _write_sentiment(wb, scan):
    ws = wb.create_sheet("Sentiment")
    labels = ["Очень негативный", "Негативный", "Нейтральный", "Позитивный", "Очень позитивный"]
    _headers(ws, ["Тип", "Разрез", *labels, "Всего"])
    _widths(ws, [16, 28, 14, 12, 12, 12, 16, 12])
    rows = [("total", "Всего", scan.get("trends", {}).get("sentiment_counts", {}))]
    rows.extend(_sentiment_breakdown(scan, "product_id", "product"))
    rows.extend(_sentiment_breakdown(scan, "source_id", "source"))
    rows.extend(_sentiment_breakdown(scan, "published_at", "month", lambda v: (v or "")[:7]))
    for kind, label, counts in rows:
        row = ws.max_row + 1
        vals = [kind, label, *[counts.get(x, 0) for x in labels], sum(counts.values())]
        for col, value in enumerate(vals, start=1):
            _cell(ws, row, col, value)


def _sentiment_breakdown(scan, field: str, kind: str, key_fn=None):
    rows = {}
    analyses = scan.get("analyses", {})
    for review in scan.get("reviews", []):
        key = review.get(field, "") or "unknown"
        if key_fn:
            key = key_fn(key)
        sentiment = analyses.get(review["review_id"], {}).get("sentiment", "Нейтральный")
        rows.setdefault(key, {})
        rows[key][sentiment] = rows[key].get(sentiment, 0) + 1
    return [(kind, key, counts) for key, counts in sorted(rows.items())]


def _write_ranked(wb, scan, sheet_name: str, key: str):
    ws = wb.create_sheet(sheet_name)
    _headers(ws, ["Ранг", "Формулировка", "Количество", "Пример", "Источник", "URL"])
    _widths(ws, [8, 55, 12, 80, 22, 45])
    for rank, item in enumerate(scan.get("insights", {}).get(key, []), start=1):
        example = item.get("example", {})
        values = [rank, item.get("item", ""), item.get("count", 0),
                  example.get("quote", ""), example.get("source", ""), example.get("url", "")]
        for col, value in enumerate(values, start=1):
            _cell(ws, rank + 1, col, value)


def _write_problems(wb, scan):
    ws = wb.create_sheet("Problems")
    _headers(ws, ["Тип", "Тема", "Отзывы", "Источники", "Негатив", "Позитив",
                  "Средний рейтинг", "Индекс критичности", "Δ к прошлому"])
    _widths(ws, [18, 32, 10, 10, 10, 10, 14, 16, 12])
    row = 2
    for kind in ("repeated_problems", "new_problems", "resolved_problems"):
        for item in scan.get("insights", {}).get(kind, []):
            values = [kind, item.get("topic", ""), item.get("reviews_count", item.get("previous_count", "")),
                      item.get("source_count", ""), item.get("negative_share", ""),
                      item.get("positive_share", ""), item.get("average_rating", ""),
                      item.get("criticality_index", ""), item.get("delta", "")]
            for col, value in enumerate(values, start=1):
                _cell(ws, row, col, value)
            row += 1


def _write_recommendations(wb, scan):
    ws = wb.create_sheet("Recommendations")
    headers = ["Priority", "Score", "Название", "Основание", "Категории",
               "Описание проблемы", "Что изменить", "Ожидаемый эффект",
               "Support", "Sources", "Quotes"]
    _headers(ws, headers)
    _widths(ws, [12, 8, 34, 58, 36, 70, 70, 65, 10, 10, 90])
    for row, sug in enumerate(scan.get("suggestions", []), start=2):
        quotes = "\n".join(
            f"{q.get('source', '')}: {q.get('quote', '')} ({q.get('url', '')})"
            for q in sug.get("quotes_with_sources", []))
        values = [sug.get("priority"), sug.get("priority_score"), sug.get("title"),
                  sug.get("basis"), "; ".join(sug.get("affected_categories", [])),
                  sug.get("problem_description"), sug.get("recommended_change"),
                  sug.get("expected_effect"), sug.get("support_count"),
                  sug.get("source_count"), quotes]
        for col, value in enumerate(values, start=1):
            _cell(ws, row, col, value)


def _write_trends(wb, scan):
    ws = wb.create_sheet("Trends")
    _headers(ws, ["Тип", "Ключ", "Значение"])
    _widths(ws, [24, 36, 18])
    row = 2
    trends = scan.get("trends", {})
    for kind in ("topic_counts", "source_counts", "product_counts", "monthly_counts"):
        for key, value in trends.get(kind, {}).items():
            for col, item in enumerate([kind, key, value], start=1):
                _cell(ws, row, col, item)
            row += 1
    for topic, metric in scan.get("insights", {}).get("topic_metrics", {}).items():
        for col, item in enumerate(["topic_metrics", topic, str(metric)], start=1):
            _cell(ws, row, col, item)
        row += 1


def _write_source_stats(wb, scan):
    ws = wb.create_sheet("Source Statistics")
    headers = ["Source ID", "Источник", "Тип", "Parser", "Policy", "Status",
               "Date Year", "Fetched", "Parsed", "Message"]
    _headers(ws, headers)
    _widths(ws, [20, 28, 18, 18, 18, 18, 10, 10, 10, 60])
    for row, stat in enumerate(scan.get("meta", {}).get("source_stats", []), start=2):
        values = [stat.get("source_id"), stat.get("source_name"), stat.get("kind"),
                  stat.get("review_parser"), stat.get("policy"), stat.get("status"),
                  stat.get("date_filter_year"), stat.get("fetched_count"),
                  stat.get("parsed_count"), stat.get("message")]
        for col, value in enumerate(values, start=1):
            _cell(ws, row, col, value)
