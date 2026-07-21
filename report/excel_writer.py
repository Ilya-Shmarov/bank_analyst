# -*- coding: utf-8 -*-
"""
Генерация Excel-отчёта (openpyxl).

Файл каждый раз регенерируется из полной истории data/history.json,
поэтому changelog и данные прошлых сканов не теряются между запусками.

Листы:
  1. Сводная            — тиры по сегментам капитала + итоговый балл + расхождения
  2. <Банк>             — детализация: значение + источник + дата проверки
  3. Lifestyle          — экосистемные подписки
  4. Изменения          — changelog (было/стало, источник, ручные уточнения)
  5. Методика оценки    — формула, веса, пороги, разбивка балла по тирам
  6. Метаданные         — дата запуска, статус источников
"""

import hashlib
import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from scanner.benefits import other_benefits_text
from scanner.formatting import (
    assert_user_visible_text,
    make_complete_summary,
    normalize_source_text,
)
from scanner.merge import field_value
from scanner.scoring import METHODOLOGY_TEXT, THRESHOLDS, WEIGHTS
from scanner.sources import (
    BANK_FIELDS,
    BANKS,
    INTL_SEGMENTS,
    LIFESTYLE_FIELDS,
    NOT_FOUND,
    NOT_FOUND_AVAILABLE,
    REFERENCE_FIELDS,
    SEGMENTS,
    SOURCE_META,
)

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
SEGMENT_FILL = PatternFill("solid", fgColor="D6E4F0")
OUR_FILL = PatternFill("solid", fgColor="E2EFDA")
DIVERGENT_FILL = PatternFill("solid", fgColor="FCE4D6")
NOT_FOUND_FONT = Font(color="999999", italic=True)
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(*[Side(style="thin", color="CCCCCC")] * 4)

LIFESTYLE_COLUMNS = dict(LIFESTYLE_FIELDS)
LIFESTYLE_EXTRA = {"bank_overlap": {"label": "Пересечения с банковскими привилегиями"}}
DISPLAY_BANK_FIELD_IDS = [
    fid for fid in BANK_FIELDS
    if fid not in {
        "taxi_restaurants",
        "always_included_options",
        "selectable_options",
        "selection_rules",
        "auto",
        "ecosystem",
    }
]


def write_report(history: dict, output_path: Path):
    wb = Workbook()
    wb.remove(wb.active)

    last_scan = history["scans"][-1] if history["scans"] else {
        "results": {}, "meta": {}, "date": ""}
    results = last_scan.get("results", {})

    _write_summary(wb, results, last_scan.get("date", ""))
    _write_bank_sheets(wb, results)
    _write_lifestyle(wb, results)
    _write_changelog(wb, history.get("changelog", []))
    _write_manual_check(wb, results)
    _write_quality_issues(wb, last_scan.get("meta", {}).get("quality_issues", []))
    _write_field_provenance(wb, results)
    _write_source_passport(wb, results)
    _write_source_conflicts(wb, results)
    _write_contract_banks(wb, results, last_scan.get("date", ""))
    _write_contract_products(wb, results)
    _write_contract_changes(wb, history.get("changelog", []))
    _write_contract_sources(wb, history)
    _write_contract_monitoring_log(wb, history)
    _write_methodology(wb, results, _last_cbr_rates(history))
    _write_meta(wb, history)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _last_cbr_rates(history: dict) -> dict:
    for scan in reversed(history.get("scans", [])):
        rates = scan.get("meta", {}).get("cbr_rates")
        if rates:
            return rates
    return {}


# ---------- helpers ----------

def _style_header_row(ws, ncols, row=1):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP
        cell.border = THIN_BORDER


def _set_widths(ws, widths):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _write_value_cell(ws, row, col, value, validate=False, context=""):
    value = _safe_excel_value(value)
    if validate and isinstance(value, str) and value:
        assert_user_visible_text(value, context or f"{ws.title}!R{row}C{col}")
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = WRAP
    cell.border = THIN_BORDER
    if value == NOT_FOUND:
        cell.font = NOT_FOUND_FONT
    return cell


def _safe_excel_value(value):
    if not isinstance(value, str):
        return value
    return ILLEGAL_CHARACTERS_RE.sub("", value)


def _display_safe_value(value):
    if value is None:
        return ""
    if isinstance(value, str) and value.strip().lower() == NOT_FOUND:
        return NOT_FOUND_AVAILABLE
    normalized = normalize_source_text(value)
    if normalized.strip().lower() == NOT_FOUND:
        return NOT_FOUND_AVAILABLE
    return normalized


def _display_url_value(value):
    if value is None:
        return ""
    url = _safe_excel_value(str(value).strip())
    url = re.sub(r"\s*:\s*//\s*", "://", url)
    return url


def _source_type_value(field: dict) -> str:
    source_type = field.get("source_type")
    if source_type:
        return source_type
    source_id = field.get("source_id", "")
    return SOURCE_META.get(source_id, {}).get("source_type", source_id)


def _annotated(field) -> str:
    """Значение поля + провенанс: источник, дата проверки, примечание."""
    if not isinstance(field, dict):
        return normalize_source_text(field) if field is not None else NOT_FOUND
    value = field.get("value", NOT_FOUND)
    if value == NOT_FOUND:
        return NOT_FOUND_AVAILABLE
    if field.get("field_id") == "other_benefits" or str(value).lstrip().startswith("•"):
        list_text = "\n".join(
            normalize_source_text(line)
            for line in str(value).splitlines()
            if line.strip()
        )
        parts = [list_text]
        src = field.get("source_name", "")
        if src:
            parts.append(f"[источник: {src}, проверено: {field.get('date_checked', '')}]")
        if field.get("note"):
            parts.append(f"[прим.: {make_complete_summary(field['note'], 180)}]")
        return "\n".join(part for part in parts if part)
    parts = [make_complete_summary(value, 420)]
    src = field.get("source_name", "")
    if src:
        parts.append(f"[источник: {src}, проверено: {field.get('date_checked', '')}]")
    if field.get("note"):
        parts.append(f"[прим.: {make_complete_summary(field['note'], 180)}]")
    return "\n".join(normalize_source_text(part) for part in parts if part)


def _conflict_status(field) -> str:
    if not isinstance(field, dict):
        return "unknown"
    if normalize_source_text(field.get("value", NOT_FOUND)) == NOT_FOUND:
        return "not_found"
    if field.get("divergent"):
        return "conflict"
    return field.get("conflict_status") or (
        "not_found" if field.get("value") == NOT_FOUND else "selected"
    )


def _reliability_status(field) -> str:
    if not isinstance(field, dict):
        return "unknown"
    if normalize_source_text(field.get("value", NOT_FOUND)) == NOT_FOUND:
        return field.get("publication_status") or "not_found"
    if field.get("divergent"):
        return "requires_manual_check"
    if field.get("publication_status") == "blocked":
        return "blocked"
    quality = field.get("quality") or "unknown"
    return f"confirmed:{quality}"


def _display_field(fields: dict, fid: str):
    if fid == "other_benefits":
        field = fields.get(fid)
        if isinstance(field, dict):
            return field
        derived = other_benefits_text(fields)
        return {"value": derived, "source_name": "Нормализация"}
    return fields.get(fid)


def _divergence_info(fields: dict) -> tuple:
    """(да/нет, комментарий по полям с расхождениями источников)."""
    comments = []
    for fid, field in fields.items():
        if isinstance(field, dict) and field.get("divergent"):
            label = BANK_FIELDS.get(fid, {}).get("label", fid)
            alt_names = ", ".join(
                a.get("source_name", "")
                for a in field.get("alternatives", [])
                if a.get("source_name")
            )
            suffix = f"; альтернативы: {alt_names}" if alt_names else ""
            comments.append(
                f"{label} — есть расхождение источников{suffix}. "
                "Детали см. на листе «Конфликты источников»."
            )
    return ("да" if comments else "нет"), "\n".join(comments)


def _tier_entries(results, types):
    for bank in BANKS:
        if bank["type"] not in types:
            continue
        for tier in bank["tiers"]:
            entry = results.get(tier["tier_id"])
            if entry:
                yield bank, tier, entry


# ---------- листы ----------

def _write_summary(wb, results, scan_date):
    ws = wb.create_sheet("Сводная")
    headers = (["Сегмент капитала", "Банк", "Тир", "Дата скана",
                "Источников OK", "Итоговый балл (0–5)",
                "Расхождение источников", "Статус достоверности",
                "Комментарий по расхождениям"]
               + [BANK_FIELDS[fid]["label"] for fid in DISPLAY_BANK_FIELD_IDS])
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, len(headers))
    _set_widths(ws, [14, 14, 24, 11, 11, 11, 13, 18, 45] + [40] * len(BANK_FIELDS))
    ws.freeze_panes = "D2"

    def write_block(segments, types, block_title=None):
        nonlocal row
        if block_title:
            cell = ws.cell(row=row, column=1, value=block_title)
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = HEADER_FILL
                ws.cell(row=row, column=col).font = HEADER_FONT
            cell.font = Font(color="FFFFFF", bold=True, size=12)
            row += 1
        for segment in segments:
            segment_row_written = False
            for bank, tier, entry in _tier_entries(results, types):
                if tier["segment"] != segment:
                    continue
                if not segment_row_written:
                    cell = ws.cell(row=row, column=1, value=segment)
                    for col in range(1, len(headers) + 1):
                        ws.cell(row=row, column=col).fill = SEGMENT_FILL
                    cell.font = Font(bold=True)
                    row += 1
                    segment_row_written = True
                divergent, div_comment = _divergence_info(entry["fields"])
                score = entry.get("score", {}).get("total", "")
                values = [
                    "", bank["name"], tier["tier_name"],
                    entry.get("scan_date", scan_date)[:10],
                    entry.get("sources_ok", ""),
                    score, divergent,
                    "requires_manual_check" if divergent == "да" else "confirmed",
                    div_comment,
                ] + [
                    _annotated(_display_field(entry["fields"], fid))
                    for fid in DISPLAY_BANK_FIELD_IDS
                ]
                for col, value in enumerate(values, start=1):
                    cell = _write_value_cell(
                        ws, row, col, value,
                        validate=True,
                        context=f"Сводная / {bank['name']} / {tier['tier_name']} / col {col}",
                    )
                    if bank["type"] == "our" and col in (2, 3):
                        cell.fill = OUR_FILL
                    if col in (7, 8) and divergent == "да":
                        cell.fill = DIVERGENT_FILL
                row += 1

    row = 2
    write_block(SEGMENTS, {"our", "bank"})
    write_block(INTL_SEGMENTS, {"intl"},
                block_title="МЕЖДУНАРОДНЫЕ DIGITAL-FIRST БАНКИ (Revolut, N26, "
                            "Wise, Monzo) — цены в валюте, баллы не считаются "
                            "(см. лист «Методика оценки»)")


def _write_bank_sheets(wb, results):
    for bank in BANKS:
        if bank["type"] not in {"our", "bank", "intl"}:
            continue
        ws = wb.create_sheet(bank["name"][:31])
        headers = ["Поле"] + [t["tier_name"] for t in bank["tiers"]]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
        _style_header_row(ws, len(headers))
        _set_widths(ws, [36] + [55] * len(bank["tiers"]))
        ws.freeze_panes = "B2"

        meta_rows = [
            ("Сегмент капитала", lambda t, e: t["segment"] or ""),
            ("Дата скана", lambda t, e: e.get("scan_date", "")[:10]),
            ("Источников OK", lambda t, e: str(e.get("sources_ok", ""))),
            ("Итоговый балл (0–5)",
             lambda t, e: str(e.get("score", {}).get("total", ""))),
            ("Источники (URL)", lambda t, e: e.get("source_url", "")),
        ]
        row = 2
        for label, getter in meta_rows:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=1).border = THIN_BORDER
            for col, tier in enumerate(bank["tiers"], start=2):
                entry = results.get(tier["tier_id"], {})
                _write_value_cell(
                    ws, row, col, getter(tier, entry) if entry else "",
                    validate=True,
                    context=f"{bank['name']} / {tier['tier_name']} / {label}",
                )
            row += 1

        for fid in DISPLAY_BANK_FIELD_IDS:
            spec = BANK_FIELDS[fid]
            ws.cell(row=row, column=1, value=spec["label"]).font = Font(bold=True)
            ws.cell(row=row, column=1).alignment = WRAP
            ws.cell(row=row, column=1).border = THIN_BORDER
            for col, tier in enumerate(bank["tiers"], start=2):
                entry = results.get(tier["tier_id"])
                field = _display_field(entry["fields"], fid) if entry else None
                cell = _write_value_cell(
                    ws, row, col, _annotated(field),
                    validate=True,
                    context=f"{bank['name']} / {tier['tier_name']} / {spec['label']}",
                )
                if isinstance(field, dict) and field.get("divergent"):
                    cell.fill = DIVERGENT_FILL
            row += 1


def _write_lifestyle(wb, results):
    ws = wb.create_sheet("Lifestyle-конкуренты")
    columns = {**LIFESTYLE_COLUMNS, **LIFESTYLE_EXTRA}
    headers = ["Подписка", "Дата скана", "Статус источника"] + [
        spec["label"] for spec in columns.values()
    ] + ["Источник"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, len(headers))
    _set_widths(ws, [22, 12, 14] + [42] * len(columns) + [40])
    ws.freeze_panes = "B2"

    row = 2
    for bank, tier, entry in _tier_entries(results, {"lifestyle"}):
        values = [
            tier["tier_name"],
            entry.get("scan_date", "")[:10],
            entry.get("status", ""),
        ] + [
            _annotated(entry["fields"].get(fid)) for fid in columns
        ] + [entry.get("source_url", "")]
        for col, value in enumerate(values, start=1):
            _write_value_cell(
                ws, row, col, value,
                validate=True,
                context=f"Lifestyle / {tier['tier_name']} / col {col}",
            )
        row += 1


def _write_changelog(wb, changelog):
    ws = wb.create_sheet("Изменения")
    headers = ["Дата скана", "Предыдущий скан", "Банк/подписка", "Тир",
               "Поле", "Было", "Стало", "Источник нового значения"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, len(headers))
    _set_widths(ws, [17, 17, 20, 22, 30, 42, 42, 30])
    ws.freeze_panes = "A2"

    if not changelog:
        ws.cell(row=2, column=1,
                value="Изменений пока нет — это первый скан или данные не менялись")
        return
    row = 2
    for change in changelog:
        if _skip_changelog_row(change):
            continue
        values = [change["scan_date"][:16], change["prev_date"][:16],
                  change["bank"], change["tier"], change["field"],
                  normalize_source_text(change["old"]),
                  normalize_source_text(change["new"]),
                  normalize_source_text(change.get("source", ""))]
        for col, value in enumerate(values, start=1):
            cell = _write_value_cell(
                ws, row, col, value,
                validate=True,
                context=f"Изменения / row {row} / col {col}",
            )
            if "ручное уточнение" in str(change.get("source", "")) and col == 8:
                cell.fill = DIVERGENT_FILL
        row += 1


def _skip_changelog_row(change: dict) -> bool:
    field = str(change.get("field", "")).lower()
    text = " ".join(str(change.get(key, "")) for key in ("old", "new", "source"))
    low = text.lower()
    if "прочее" in field and ("рейтинг уровня по отзывам" in low or "отзыв" in low):
        return True
    return False


def _write_manual_check(wb, results):
    """Целевые поля со статусом «не найдено» — чтобы пробелы не терялись молча.
    Значения «—» (подтверждённое отсутствие услуги) сюда не попадают."""
    ws = wb.create_sheet("Требует ручной проверки")
    headers = ["Банк/подписка", "Тир", "Поле", "Причина / что делать"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, len(headers))
    _set_widths(ws, [24, 26, 38, 70])
    ws.freeze_panes = "A2"

    all_labels = {**{k: v["label"] for k, v in BANK_FIELDS.items()},
                  **{k: v["label"] for k, v in LIFESTYLE_FIELDS.items()},
                  "bank_overlap": "Пересечения с банковскими привилегиями"}
    row = 2
    for bank in BANKS:
        for tier in bank["tiers"]:
            entry = results.get(tier["tier_id"])
            if not entry:
                continue
            for fid, field in entry["fields"].items():
                if fid in REFERENCE_FIELDS:
                    continue
                if field_value(field) != NOT_FOUND:
                    continue
                if isinstance(field, dict) and field.get("publication_status") == "blocked":
                    reason = (
                        "значение заблокировано для HTML: "
                        f"{field.get('publication_reason', '')}; "
                        f"исходное значение: {_safe_reason_text(field.get('blocked_value', ''))}"
                    )
                elif entry.get("sources_ok", 0) == 0:
                    reason = ("все источники недоступны (антибот/блокировка) — "
                              "проверить вручную или найти зеркальный источник")
                else:
                    reason = ("целевой поиск публичных данных результата не дал — "
                              "проверить тарифные PDF банка / запросить у банка")
                values = [entry["bank"], entry["tier"], all_labels.get(fid, fid),
                          reason]
                for col, value in enumerate(values, start=1):
                    _write_value_cell(
                        ws, row, col, value,
                        validate=True,
                        context=f"Требует ручной проверки / {entry['bank']} / {entry['tier']} / col {col}",
                    )
                row += 1
    if row == 2:
        ws.cell(row=2, column=1, value="Пробелов нет — все поля заполнены "
                                       "или помечены как отсутствующие")


def _safe_reason_text(value) -> str:
    return make_complete_summary(
        normalize_source_text(str(value).replace("|", ";")),
        260,
    )


def _write_field_provenance(wb, results):
    """Служебная трассировка: откуда взято каждое итоговое значение."""
    ws = wb.create_sheet("Провенанс значений")
    headers = [
        "bank_id", "Банк", "tier_id", "Тир", "field_id", "Поле",
        "Итоговое значение", "source_url", "source_section", "source_type", "checked_at",
        "raw_text", "status", "conflict_status", "publication_status",
        "publication_reason", "blocked_value", "reason",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, len(headers))
    _set_widths(ws, [18, 20, 26, 30, 24, 34, 54, 60, 30, 16, 14, 60, 24, 18, 20, 52, 54, 42])
    ws.freeze_panes = "A2"

    labels = {**{k: v["label"] for k, v in BANK_FIELDS.items()},
              **{k: v["label"] for k, v in LIFESTYLE_FIELDS.items()},
              "bank_overlap": "Пересечения с банковскими привилегиями"}
    row = 2
    for bank in BANKS:
        for tier in bank["tiers"]:
            entry = results.get(tier["tier_id"])
            if not entry:
                continue
            for fid, field in entry.get("fields", {}).items():
                if fid in REFERENCE_FIELDS:
                    continue
                values = [
                    bank["id"], bank["name"], tier["tier_id"], tier["tier_name"],
                    fid, labels.get(fid, fid), _display_safe_value(field_value(field)),
                    _display_url_value(field.get("source_url", "")) if isinstance(field, dict) else "",
                    _display_safe_value(field.get("source_section", "") or tier["tier_name"]) if isinstance(field, dict) else "",
                    _display_safe_value(_source_type_value(field)) if isinstance(field, dict) else "",
                    _display_safe_value(field.get("date_checked", "")) if isinstance(field, dict) else "",
                    _display_safe_value(field.get("raw_text", field_value(field))) if isinstance(field, dict) else _display_safe_value(field_value(field)),
                    _display_safe_value(_reliability_status(field)),
                    _display_safe_value(_conflict_status(field)),
                    _display_safe_value(field.get("publication_status", "")) if isinstance(field, dict) else "",
                    _display_safe_value(field.get("publication_reason", "")) if isinstance(field, dict) else "",
                    _display_safe_value(field.get("blocked_value", "")) if isinstance(field, dict) else "",
                    _display_safe_value(field.get("note", "")) if isinstance(field, dict) else "",
                ]
                for col, value in enumerate(values, start=1):
                    cell = _write_value_cell(ws, row, col, value)
                    if col == 14 and value == "conflict":
                        cell.fill = DIVERGENT_FILL
                    if col == 15 and value == "blocked":
                        cell.fill = DIVERGENT_FILL
                row += 1


def _write_source_passport(wb, results):
    """Per-tier publication passport for management-facing facts."""
    ws = wb.create_sheet("Source_Passport")
    headers = [
        "bank_id", "Банк", "tier_id", "Тир", "configured_urls",
        "published_fields", "blocked_fields", "not_found_fields",
        "curated_fields", "pbi_fallback_fields", "official_fields",
        "blocked_reasons",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, len(headers))
    _set_widths(ws, [16, 20, 26, 30, 70, 44, 44, 44, 44, 44, 44, 80])
    ws.freeze_panes = "A2"

    row = 2
    labels = {k: v["label"] for k, v in BANK_FIELDS.items()}
    for bank in BANKS:
        if bank["type"] not in {"our", "bank", "intl"}:
            continue
        for tier in bank["tiers"]:
            entry = results.get(tier["tier_id"])
            if not entry:
                continue
            fields = {
                fid: field
                for fid, field in entry.get("fields", {}).items()
                if fid in DISPLAY_BANK_FIELD_IDS
            }
            values = [
                bank["id"],
                bank["name"],
                tier["tier_id"],
                tier["tier_name"],
                _configured_urls(tier),
                _fields_by_publication_status(fields, "published", labels),
                _fields_by_publication_status(fields, "blocked", labels),
                _fields_by_publication_status(fields, "not_found", labels),
                _fields_by_source_id(fields, "curated", labels),
                _fields_by_source_type(fields, "premiumbanking.info", labels),
                _fields_by_source_type(fields, "official", labels),
                _blocked_reasons(fields, labels),
            ]
            for col, value in enumerate(values, start=1):
                _write_value_cell(ws, row, col, _display_safe_value(value))
            row += 1


def _configured_urls(tier: dict) -> str:
    urls = []
    for src in tier.get("sources", []):
        urls.extend(src.get("urls", []))
    return "\n".join(dict.fromkeys(urls))


def _fields_by_publication_status(fields: dict, status: str, labels: dict) -> str:
    names = []
    for fid, field in fields.items():
        if not isinstance(field, dict):
            continue
        if field.get("publication_status") == status:
            names.append(labels.get(fid, fid))
    return "\n".join(names)


def _fields_by_source_id(fields: dict, source_id: str, labels: dict) -> str:
    names = []
    for fid, field in fields.items():
        if not isinstance(field, dict):
            continue
        if field.get("publication_status") == "published" and field.get("source_id") == source_id:
            names.append(labels.get(fid, fid))
    return "\n".join(names)


def _fields_by_source_type(fields: dict, source_type: str, labels: dict) -> str:
    names = []
    for fid, field in fields.items():
        if not isinstance(field, dict):
            continue
        if (field.get("publication_status") == "published"
                and _source_type_value(field) == source_type):
            names.append(labels.get(fid, fid))
    return "\n".join(names)


def _blocked_reasons(fields: dict, labels: dict) -> str:
    rows = []
    for fid, field in fields.items():
        if not isinstance(field, dict):
            continue
        if field.get("publication_status") == "blocked":
            rows.append(f"{labels.get(fid, fid)}: {field.get('publication_reason', '')}")
    return "\n".join(rows)


def _write_source_conflicts(wb, results):
    """Отдельный лист для полей, где источники дали разные числовые значения."""
    ws = wb.create_sheet("Конфликты источников")
    headers = [
        "bank_id", "Банк", "tier_id", "Тир", "field_id", "Поле",
        "Официальное значение", "Официальный URL",
        "Значение PremiumBanking.info", "URL PremiumBanking.info",
        "Выбранное итоговое значение", "Причина выбора", "Дата проверки",
        "Выбранный тип", "Альтернативный источник",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, len(headers))
    _set_widths(ws, [18, 20, 26, 30, 24, 34, 54, 60, 54, 60, 54, 48, 16, 20, 28])
    ws.freeze_panes = "A2"

    labels = {k: v["label"] for k, v in BANK_FIELDS.items()}
    row = 2
    for bank, tier, entry in _tier_entries(results, {"our", "bank", "intl"}):
        for fid, field in entry.get("fields", {}).items():
            if fid in REFERENCE_FIELDS or not isinstance(field, dict):
                continue
            if not field.get("divergent"):
                continue
            alternatives = field.get("alternatives") or [{}]
            for alt in alternatives:
                official_value = ""
                official_url = ""
                pbi_value = ""
                pbi_url = ""
                selected_type = _source_type_value(field)
                alt_type = _source_type_value(alt)
                if selected_type == "official":
                    official_value = field.get("value", NOT_FOUND)
                    official_url = field.get("source_url", "")
                if alt_type == "official":
                    official_value = alt.get("value", "")
                    official_url = alt.get("source_url", "")
                if selected_type == "premiumbanking.info":
                    pbi_value = field.get("value", NOT_FOUND)
                    pbi_url = field.get("source_url", "")
                if alt_type == "premiumbanking.info":
                    pbi_value = alt.get("value", "")
                    pbi_url = alt.get("source_url", "")
                values = [
                    bank["id"], bank["name"], tier["tier_id"], tier["tier_name"],
                    fid, labels.get(fid, fid),
                    _display_safe_value(official_value),
                    _display_safe_value(official_url),
                    _display_safe_value(pbi_value),
                    _display_safe_value(pbi_url),
                    _display_safe_value(field.get("value", NOT_FOUND)),
                    _display_safe_value(field.get("note", "")),
                    _display_safe_value(field.get("date_checked", "")),
                    _display_safe_value(selected_type),
                    _display_safe_value(alt.get("source_name", "")),
                ]
                for col, value in enumerate(values, start=1):
                    _write_value_cell(ws, row, col, value)
                row += 1
    if row == 2:
        ws.cell(row=2, column=1, value="Конфликтов источников нет")


def _write_quality_issues(wb, issues):
    ws = wb.create_sheet("Quality_Issues")
    headers = [
        "severity", "code", "bank", "tier", "tier_id",
        "field_id", "message", "value",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, len(headers))
    _set_widths(ws, [12, 22, 22, 34, 26, 24, 48, 70])
    ws.freeze_panes = "A2"

    row = 2
    for issue in issues:
        values = [
            issue.get("severity", ""),
            issue.get("code", ""),
            issue.get("bank", ""),
            issue.get("tier", ""),
            issue.get("tier_id", ""),
            issue.get("field_id", ""),
            issue.get("message", ""),
            issue.get("value", ""),
        ]
        for col, value in enumerate(values, start=1):
            cell = _write_value_cell(ws, row, col, _display_safe_value(value))
            if col == 1 and value == "error":
                cell.fill = DIVERGENT_FILL
        row += 1
    if row == 2:
        ws.cell(row=2, column=1, value="Ошибок качества нет")


# ---------- contract sheets requested by the monitoring process ----------

CHANGE_CATEGORIES = {
    "Стоимость обслуживания": "Стоимость обслуживания",
    "Условия входа / поддержания уровня": "Условия бесплатности",
    "Бизнес-залы (визиты, спутники)": "Бизнес-залы",
    "Страхование": "Страхование",
    "Консьерж-сервис": "Консьерж-сервис",
    "Кэшбэк (ставка, категории, механика)": "Кешбэк",
    "Карты (тип, лимиты переводов/снятия, выпуск)": "Карты",
    "Такси": "Партнерские предложения",
    "Рестораны": "Партнерские предложения",
    "Вклады / накопительные счета": "Вклады",
    "Спецусловия по вкладам / накопительным счетам": "Вклады",
    "Другие привилегии": "Состав пакета",
    "Экосистемные привилегии (доставка, подписки)": "Состав пакета",
}


def _write_contract_banks(wb, results, scan_date):
    ws = wb.create_sheet("Banks")
    headers = [
        "bank_id", "bank_name", "bank_slug", "official_site",
        "premium_page_url", "news_page_url", "tariffs_page_url",
        "monitoring_status", "last_checked_at",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, len(headers))
    _set_widths(ws, [18, 24, 22, 42, 58, 42, 58, 18, 20])

    row = 2
    for bank in BANKS:
        if bank["type"] == "lifestyle":
            continue
        source_urls = [
            url
            for tier in bank["tiers"]
            for src in tier.get("sources", [])
            for url in src.get("urls", [])
        ]
        official = next((url for url in source_urls if "premiumbanking.info" not in url), "")
        pbi = next((url for url in source_urls if "premiumbanking.info" in url), "")
        status = "active" if any(t["tier_id"] in results for t in bank["tiers"]) else "configured"
        values = [
            bank["id"], bank["name"], bank["id"], official, official or pbi,
            "", official, status, scan_date,
        ]
        for col, value in enumerate(values, start=1):
            _write_value_cell(ws, row, col, _display_safe_value(value))
        row += 1


def _write_contract_products(wb, results):
    ws = wb.create_sheet("Products")
    headers = [
        "bank_id", "product_id", "product_name", "product_type",
        "service_cost", "free_service_conditions", "minimum_balance",
        "required_turnover", "required_income", "lounge_access", "lounge_limit",
        "travel_insurance", "concierge", "cashback", "family_privileges",
        "cash_withdrawal", "transfers", "investments", "deposits",
        "additional_privileges", "effective_from", "source_url",
        "source_checked_at", "status",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, len(headers))
    _set_widths(ws, [16, 26, 30, 18] + [34] * (len(headers) - 4))
    ws.freeze_panes = "A2"

    row = 2
    for bank in BANKS:
        if bank["type"] == "lifestyle":
            continue
        for tier in bank["tiers"]:
            entry = results.get(tier["tier_id"])
            fields = entry.get("fields", {}) if entry else {}
            values = [
                bank["id"],
                tier["tier_id"],
                tier["tier_name"],
                bank["type"],
                _contract_field(fields, "service_cost"),
                _contract_field(fields, "entry_conditions"),
                NOT_FOUND,
                NOT_FOUND,
                NOT_FOUND,
                _contract_field(fields, "lounge_access"),
                NOT_FOUND,
                _contract_field(fields, "insurance"),
                _contract_field(fields, "concierge"),
                _contract_field(fields, "cashback"),
                NOT_FOUND,
                _contract_field(fields, "card_terms"),
                _contract_field(fields, "card_terms"),
                NOT_FOUND,
                _contract_field(fields, "deposits"),
                _contract_field(_display_field(fields, "other_benefits"), None),
                _contract_field(fields, "last_change_date"),
                entry.get("source_url", "") if entry else "",
                entry.get("scan_date", "") if entry else "",
                entry.get("status", "not_scanned") if entry else "not_scanned",
            ]
            for col, value in enumerate(values, start=1):
                _write_value_cell(ws, row, col, _display_safe_value(value))
            row += 1


def _contract_field(fields_or_field, field_id):
    if field_id is None:
        return field_value(fields_or_field)
    return field_value(fields_or_field.get(field_id, {"value": NOT_FOUND}))


def _write_contract_changes(wb, changelog):
    ws = wb.create_sheet("Changes")
    headers = [
        "change_id", "bank_id", "bank_name", "product_id", "product_name",
        "category", "change_type", "old_value", "new_value",
        "short_description", "full_description", "published_at",
        "effective_from", "detected_at", "source_title", "source_url",
        "additional_sources", "source_type", "source_quote",
        "document_version", "content_hash", "confidence",
        "verification_status", "is_temporary", "valid_until",
        "analyst_comment", "html_visible", "record_status",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, len(headers))
    _set_widths(ws, [32, 16, 24, 26, 30, 24, 16, 42, 42, 42, 60, 16, 16, 18,
                     28, 58, 40, 20, 50, 24, 24, 14, 26, 12, 16, 44, 12, 16])
    ws.freeze_panes = "A2"

    row = 2
    seen = set()
    for change in changelog:
        if change.get("kind") != "market":
            continue
        key = (
            change.get("bank", ""),
            change.get("tier", ""),
            change.get("field", ""),
            normalize_source_text(change.get("new", "")),
            change.get("source_url", ""),
        )
        if key in seen:
            continue
        seen.add(key)
        bank_id = _bank_id_by_name(change.get("bank", ""))
        product_id = _tier_id_by_name(change.get("tier", ""))
        detected_at = change.get("scan_date", "")
        category = CHANGE_CATEGORIES.get(change.get("field", ""), "Прочее")
        source_url = _display_url_value(change.get("source_url", ""))
        verification = "Подтверждено" if source_url else "Требует ручной проверки"
        confidence = _confidence_from_url(source_url)
        change_id = _change_id(bank_id, product_id, category, detected_at, change.get("new", ""))
        short = f"{change.get('field', 'Условия')} изменено"
        full = (
            f"{change.get('bank', '')} / {change.get('tier', '')}: "
            f"{change.get('field', '')} изменено с "
            f"{normalize_source_text(change.get('old', ''))} на "
            f"{normalize_source_text(change.get('new', ''))}"
        )
        values = [
            change_id, bank_id, change.get("bank", ""), product_id,
            change.get("tier", ""), category, _change_type(change),
            normalize_source_text(change.get("old", "")),
            normalize_source_text(change.get("new", "")),
            short, full, "не указано", "не указано", detected_at,
            change.get("source", ""), source_url, "", _source_type_from_url(source_url),
            normalize_source_text(change.get("new", "")), "", _content_hash(change),
            confidence, verification, "false", "", "", "true", "Актуальная",
        ]
        url_columns = {"source_url", "additional_sources"}
        for col, value in enumerate(values, start=1):
            header = headers[col - 1]
            safe_value = _display_url_value(value) if header in url_columns else _display_safe_value(value)
            _write_value_cell(ws, row, col, safe_value)
        row += 1
    if row == 2:
        ws.cell(row=2, column=1, value="Подтвержденных рыночных изменений пока нет")


def _write_contract_sources(wb, history):
    ws = wb.create_sheet("Sources")
    headers = [
        "source_id", "bank_id", "url", "page_type", "title", "http_status",
        "checked_at", "content_hash", "previous_content_hash",
        "content_changed", "parse_status", "error_message",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, len(headers))
    _set_widths(ws, [30, 16, 70, 18, 30, 14, 18, 24, 24, 16, 18, 44])
    ws.freeze_panes = "A2"

    last_scan = history.get("scans", [{}])[-1] if history.get("scans") else {}
    checked_at = last_scan.get("date", "")
    failed = last_scan.get("meta", {}).get("sources_failed", {})
    row = 2
    for bank in BANKS:
        for tier in bank["tiers"]:
            for src in tier.get("sources", []):
                for idx, url in enumerate(src.get("urls", []), start=1):
                    title = f"{bank['name']} / {tier['tier_name']} [{src['source_id']}]"
                    error = failed.get(title, "")
                    values = [
                        f"{tier['tier_id']}__{src['source_id']}__{idx}",
                        bank["id"], url, src["source_id"], title,
                        _http_status_from_error(error), checked_at, "", "",
                        "", "unavailable" if error else "configured", error,
                    ]
                    for col, value in enumerate(values, start=1):
                        safe_value = _display_url_value(value) if headers[col - 1] == "url" else _display_safe_value(value)
                        _write_value_cell(ws, row, col, safe_value)
                    row += 1


def _write_contract_monitoring_log(wb, history):
    ws = wb.create_sheet("Monitoring_Log")
    headers = [
        "run_id", "started_at", "finished_at", "bank_id", "pages_checked",
        "documents_checked", "changes_detected", "changes_confirmed",
        "errors", "run_status",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, len(headers))
    _set_widths(ws, [28, 20, 20, 16, 16, 18, 18, 18, 12, 18])
    ws.freeze_panes = "A2"

    for row, scan in enumerate(history.get("scans", []), start=2):
        meta = scan.get("meta", {})
        failed = meta.get("sources_failed", {})
        changes = meta.get("changes_found", 0)
        values = [
            _run_id(scan.get("date", "")), scan.get("date", ""), scan.get("date", ""),
            meta.get("mode", "all"), len(meta.get("sources_ok", [])),
            0, changes, changes, len(failed), "ok" if not failed else "partial",
        ]
        for col, value in enumerate(values, start=1):
            _write_value_cell(ws, row, col, _display_safe_value(value))


def _bank_id_by_name(name):
    for bank in BANKS:
        if bank["name"] == name:
            return bank["id"]
    return ""


def _tier_id_by_name(name):
    for bank in BANKS:
        for tier in bank["tiers"]:
            if tier["tier_name"] == name:
                return tier["tier_id"]
    return ""


def _change_type(change):
    old = normalize_source_text(change.get("old", ""))
    new = normalize_source_text(change.get("new", ""))
    if old == NOT_FOUND and new != NOT_FOUND:
        return "Добавлено"
    if new == NOT_FOUND:
        return "Удалено"
    return "Изменено"


def _source_type_from_url(url):
    if not url:
        return ""
    if "premiumbanking.info" in url:
        return "premiumbanking.info"
    return "official"


def _confidence_from_url(url):
    source_type = _source_type_from_url(url)
    if source_type == "official":
        return "Высокий"
    if source_type == "premiumbanking.info":
        return "Низкий"
    return "Низкий" if url else ""


def _content_hash(value):
    data = json.dumps(value, ensure_ascii=False, sort_keys=True)
    return hashlib.sha256(data.encode("utf-8")).hexdigest()[:16]


def _change_id(bank_id, product_id, category, detected_at, new_value):
    date = (detected_at or "")[:10].replace("-", "") or "unknown"
    raw = "|".join([bank_id, product_id, category, detected_at or "", str(new_value)])
    digest = hashlib.sha1(raw.encode("utf-8")).hexdigest()[:8]
    base = "_".join(part for part in [bank_id, product_id, category, date] if part)
    base = "".join(ch if ch.isalnum() else "_" for ch in base).strip("_")
    return f"{base}_{digest}" if base else digest


def _run_id(scan_date):
    return f"run_{(scan_date or '').replace('-', '').replace(':', '').replace('T', '_')}"


def _http_status_from_error(error):
    if not error:
        return ""
    for token in str(error).replace(":", " ").split():
        if token.isdigit() and 100 <= int(token) <= 599:
            return token
    return ""


INTL_METHODOLOGY_TEXT = [
    "",
    "МЕЖДУНАРОДНЫЕ DIGITAL-FIRST БАНКИ (Revolut, N26, Wise, Monzo) — "
    "методологические оговорки:",
    "1. Цены подписок и лимиты в валюте страны (£/€) НЕЛЬЗЯ сравнивать 1:1 "
    "с рублёвыми пакетами: без поправки на масштаб экономики и доходы "
    "сегмента прямое сопоставление вводит в заблуждение. Блок — для "
    "сравнения продуктовых механик (подписочная модель, страховые пакеты, "
    "metal-карты, перки), а не абсолютных цифр.",
    "2. По этой же причине международный блок НЕ участвует в балльной "
    "оценке (пороговые таблицы откалиброваны под рублёвый рынок).",
    "3. Прочерк «— (не предусмотрено моделью продукта)» — подтверждённое "
    "отсутствие категории у необанка (например, консьерж или автоуслуги), "
    "а не пробел данных.",
    "4. Wise включён без деления на тиры — как конкурент за multi-currency "
    "lifestyle-сценарий (модель pay-per-use без подписки).",
    "5. Цены в оригинальной валюте; для справочного пересчёта в рубли — "
    "курсы ЦБ на дату скана (строка ниже).",
]


def _write_methodology(wb, results, cbr_rates=None):
    ws = wb.create_sheet("Методика оценки")
    _set_widths(ws, [26, 22, 34, 10, 9, 13])

    row = 1
    ws.cell(row=row, column=1, value="Методика собственной оценки пакетов")
    ws.cell(row=row, column=1).font = Font(bold=True, size=13)
    row += 2
    for line in METHODOLOGY_TEXT + INTL_METHODOLOGY_TEXT:
        cell = ws.cell(row=row, column=1, value=line)
        cell.alignment = WRAP
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1
    if cbr_rates:
        date = cbr_rates.get("date", "")
        rates_line = "Курсы ЦБ РФ на " + date + ": " + "; ".join(
            f"{code} = {value} ₽" for code, value in cbr_rates.items()
            if code != "date")
        cell = ws.cell(row=row, column=1, value=rates_line)
        cell.font = Font(bold=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="Веса категорий (сумма = 1.0)").font = Font(bold=True)
    row += 1
    for category, weight in WEIGHTS.items():
        label = BANK_FIELDS.get(category, {}).get("label", category)
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=weight)
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="Пороговые таблицы (метрика ≥ порога → балл)").font = Font(bold=True)
    row += 1
    for table_key, table in THRESHOLDS.items():
        ws.cell(row=row, column=1, value=table_key)
        ws.cell(row=row, column=2,
                value="; ".join(f"≥{minimum} → {score}" for minimum, score in table))
        row += 1
    ws.cell(row=row, column=1, value="особые правила")
    ws.cell(row=row, column=2,
            value="безлимит → 5; консьерж есть/нет → 5/0; страховка по покрытию "
                  "(≈1 млн → 5 … 30–90 тыс → 2); авто: 3 базово +1 помощь на "
                  "дорогах +1 кэшбэк дороги/парковки; «не найдено» → 0")
    ws.cell(row=row, column=2).alignment = WRAP
    row += 2

    ws.cell(row=row, column=1, value="Разбивка балла по тирам").font = Font(bold=True)
    row += 1
    headers = ["Банк / тир", "Категория", "Извлечённая метрика", "Балл 0–5",
               "Вес", "Вклад в итог"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=row, column=col, value=header)
    _style_header_row(ws, len(headers), row=row)
    row += 1

    for bank, tier, entry in _tier_entries(results, {"our", "bank"}):
        score = entry.get("score")
        if not score:
            continue
        first_row = row
        for category, detail in score["breakdown"].items():
            label = BANK_FIELDS.get(category, {}).get("label", category)
            values = ["", label, detail["metric"], detail["score"],
                      detail["weight"], detail["contribution"]]
            for col, value in enumerate(values, start=1):
                _write_value_cell(ws, row, col, value)
            row += 1
        name_cell = ws.cell(row=first_row, column=1,
                            value=f"{bank['name']} / {tier['tier_name']}")
        name_cell.font = Font(bold=True)
        name_cell.alignment = WRAP
        total_cell = ws.cell(row=row, column=1, value="ИТОГО")
        total_cell.font = Font(bold=True)
        ws.cell(row=row, column=6, value=score["total"]).font = Font(bold=True)
        row += 2


def _write_meta(wb, history):
    ws = wb.create_sheet("Метаданные")
    _set_widths(ws, [30, 110])
    ws.cell(row=1, column=1, value="Параметр")
    ws.cell(row=1, column=2, value="Значение")
    _style_header_row(ws, 2)

    row = 2
    for scan in reversed(history.get("scans", [])):
        meta = scan.get("meta", {})
        rows = [
            ("Дата запуска", scan.get("date", "")),
            ("Режим", meta.get("mode", "")),
            ("Источники OK", "; ".join(meta.get("sources_ok", [])) or "—"),
            ("Источники с ошибками",
             "; ".join(f"{k}: {v}" for k, v in meta.get("sources_failed", {}).items()) or "—"),
            ("Полей обновлено", str(meta.get("fields_updated", ""))),
            ("Изменений найдено", str(meta.get("changes_found", ""))),
        ]
        for label, value in rows:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            _write_value_cell(ws, row, 2, value)
            row += 1
        row += 1  # пустая строка между сканами
